# -*- coding: utf-8 -*-
"""Append new jobs to the tracker (creates a styled jobs.xlsx if missing).

Usage:
    python add_jobs.py init                 # just create/ensure the workbook
    python add_jobs.py <new_jobs.json>      # append jobs, dedupe by link/job id

new_jobs.json = [
  {"company": "...", "title": "...", "location": "...", "link": "https://...",
   "job_id": "...", "score": 8, "reason": "...", "date": "2026-07-30"}, ...
]
Prints a summary; exits 0 on success. No formulas are used in this workbook.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TRACKER = Path(__file__).resolve().parent.parent / "jobs.xlsx"
SHEET = "Jobs"

HEADERS = ["#", "תאריך", "חברה", "משרה", "מיקום", "ציון", "למה מתאים",
           "קישור", "Job ID", "סטטוס", "קובץ CV", "הערות", "הוגש בתאריך"]
WIDTHS = [5, 11, 14, 38, 16, 6, 52, 14, 14, 10, 14, 20, 12]
STATUS_LIST = '"חדש,הוגש,ראיון,נדחה,דילגתי"'

HEADER_FILL = PatternFill("solid", fgColor="16385E")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="2563A8", underline="single")


def ensure_workbook():
    if TRACKER.exists():
        return load_workbook(TRACKER)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    for col, (head, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=head)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    dv = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("J2:J1000")

    legend = wb.create_sheet("Legend")
    legend.sheet_view.rightToLeft = True
    legend.column_dimensions["A"].width = 16
    legend.column_dimensions["B"].width = 80
    rows = [
        ("מה זה?", "טראקר משרות - מתמלא אוטומטית ע\"י הסריקה השבועית של Claude."),
        ("עמודות לעריכה", "סטטוס (רשימה נפתחת) והערות - שלך. שאר העמודות מנוהלות אוטומטית."),
        ("סטטוסים", "חדש = טרם טופל | הוגש = שלחת מועמדות | ראיון | נדחה | דילגתי"),
        ("יצירת CV", "בדשבורד: כפתור Generate CV מעתיק פקודה - מדביקים בצ'אט של Claude."),
        ("ציון", "התאמה 1-10 לפי profile.md (סף כניסה: 6)."),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        legend.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        legend.cell(row=r, column=2, value=b).font = BODY_FONT
    wb.save(TRACKER)
    return load_workbook(TRACKER)


def existing_keys(wb):
    """Dedupe keys from live rows, the Deleted tombstone sheet AND the Archive
    sheet - a job the user deleted or that went stale must never be re-added."""
    keys = set()
    ws = wb[SHEET]
    for row in ws.iter_rows(min_row=2):
        link = row[7].hyperlink.target if row[7].hyperlink else row[7].value
        if link and str(link).startswith("http"):
            keys.add(str(link).strip().rstrip("/").lower())     # real URL, not display text
        if row[8].value:
            keys.add(str(row[8].value).strip().lower())         # job id
    if "Deleted" in wb.sheetnames:
        for row in wb["Deleted"].iter_rows(min_row=2, values_only=True):
            if len(row) > 3 and row[3]:
                keys.add(str(row[3]).strip().rstrip("/").lower())
            if len(row) > 4 and row[4]:
                keys.add(str(row[4]).strip().lower())
    if "Archive" in wb.sheetnames:  # same column layout as Jobs
        for row in wb["Archive"].iter_rows(min_row=2, values_only=True):
            if len(row) > 7 and row[7] and str(row[7]).startswith("http"):
                keys.add(str(row[7]).strip().rstrip("/").lower())
            if len(row) > 8 and row[8]:
                keys.add(str(row[8]).strip().lower())
    return keys


def next_id(wb):
    """Max id across Jobs AND Archive - an archived row must keep its id
    reserved, or a new job could collide with its CV files in output/."""
    ids = [row[0] for row in wb[SHEET].iter_rows(min_row=2, values_only=True)
           if isinstance(row[0], int)]
    if "Archive" in wb.sheetnames:
        ids += [row[0] for row in wb["Archive"].iter_rows(min_row=2, values_only=True)
                if isinstance(row[0], int)]
    return (max(ids) + 1) if ids else 1


def last_data_row(ws):
    """Last row that actually holds data - ignores styled-but-empty ghost rows."""
    for r in range(ws.max_row, 1, -1):
        if any(c.value is not None for c in ws[r]):
            return r
    return 1


def add_jobs(jobs):
    wb = ensure_workbook()
    ws = wb[SHEET]
    seen = existing_keys(wb)
    nid = next_id(wb)
    added, skipped = 0, 0
    for j in jobs:
        link_key = str(j.get("link", "")).strip().rstrip("/").lower()
        id_key = str(j.get("job_id", "")).strip().lower()
        if (link_key and link_key in seen) or (id_key and id_key in seen):
            skipped += 1
            continue
        row = last_data_row(ws) + 1
        values = [nid, j.get("date", ""), j.get("company", ""), j.get("title", ""),
                  j.get("location", ""), j.get("score", ""), j.get("reason", ""),
                  "קישור", j.get("job_id", ""), "חדש", "", ""]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (4, 7)))
        # plain URL as the cell VALUE - openpyxl hyperlink objects break on delete_rows
        link_cell = ws.cell(row=row, column=8)
        link_cell.value = j.get("link", "")
        link_cell.font = LINK_FONT
        if link_key:
            seen.add(link_key)
        if id_key:
            seen.add(id_key)
        nid += 1
        added += 1
    wb.save(TRACKER)
    print(f"added={added} skipped_duplicates={skipped} total_rows={last_data_row(ws) - 1}")


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    if sys.argv[1] == "init":
        ensure_workbook()
        print(f"tracker ready: {TRACKER}")
        return
    jobs = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8-sig"))
    add_jobs(jobs)


if __name__ == "__main__":
    main()
