# -*- coding: utf-8 -*-
"""Auto-archive stale tracker rows (jobs.xlsx -> "Archive" sheet).

Policy (age = today minus the date column, i.e. posting date when known):
  - "חדש" older than 7 days            -> archived (Lidor 2026-09-02: a week
                                          without applying = not applying)
  - "נדחה"/"דילגתי" older than 30 days -> archived (declutter, kept for history)
  - "הוגש"/"ראיון" are NEVER auto-archived.
  - Archive rows older than 7 days (by archive date) -> DELETED for good
    (Lidor 2026-09-02), except rows that were ever הוגש/ראיון. A tombstone
    goes to the "Deleted" sheet so scans never re-add them.
Archived rows keep every column + archive date + reason, stay in the dedupe set
(add_jobs.existing_keys reads Archive too) so scans never re-add them, and can
be restored by hand (ask Claude) while they are still in the archive.

Usage: python archive_stale.py [--dry-run]
Also runs automatically at the start of every make_dashboard.py rebuild.
"""
import datetime
import sys
from pathlib import Path

from openpyxl import load_workbook

TRACKER = Path(__file__).resolve().parent.parent / "jobs.xlsx"
STALE_NEW_DAYS = 7      # status "חדש" (was 21; Lidor 2026-09-02)
STALE_CLOSED_DAYS = 30  # status "נדחה" / "דילגתי"
PURGE_ARCHIVE_DAYS = 7  # archived this long -> deleted for good (Lidor 2026-09-02)
ARCHIVE_SHEET = "Archive"
DELETED_SHEET = "Deleted"
SUBMIT_DATE_COL = 13    # Jobs col "הוגש בתאריך" (cv_server writes it on ✅)

ARCHIVE_HEADERS = ["#", "תאריך", "חברה", "משרה", "מיקום", "ציון", "למה מתאים",
                   "קישור", "Job ID", "סטטוס", "קובץ CV", "הערות",
                   "הוגש בתאריך", "תאריך ארכוב", "סיבה"]


def parse_date(v):
    try:
        return datetime.date.fromisoformat(str(v)[:10])
    except (ValueError, TypeError):
        return None


def ensure_schema(wb):
    """Archive sheet + the submit-date column on Jobs. Returns (sheet, changed)."""
    changed = False
    ws = wb["Jobs"]
    if ws.cell(row=1, column=SUBMIT_DATE_COL).value is None:
        ws.cell(row=1, column=SUBMIT_DATE_COL, value="הוגש בתאריך")
        ws.column_dimensions["M"].width = 12
        changed = True
    if ARCHIVE_SHEET not in wb.sheetnames:
        ar = wb.create_sheet(ARCHIVE_SHEET)
        ar.sheet_view.rightToLeft = True
        for col, head in enumerate(ARCHIVE_HEADERS, start=1):
            ar.cell(row=1, column=col, value=head)
        changed = True
    return wb[ARCHIVE_SHEET], changed


def run(dry_run=False, today=None):
    """Move stale rows to Archive. Returns (count, ["#16 ISCAR (reason)", ...])."""
    today = today or datetime.date.today()
    wb = load_workbook(TRACKER)
    ws = wb["Jobs"]
    ar, schema_changed = ensure_schema(wb)
    stale = []  # (sheet_row_idx, values, reason)
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        if "שוחזר" in str(row[11].value or ""):
            continue  # user pulled it back from the archive on purpose - keep it
        status = row[9].value or "חדש"
        d = parse_date(row[1].value)
        if d is None:
            continue  # unparseable date: never auto-archive
        age = (today - d).days
        reason = None
        if status == "חדש" and age >= STALE_NEW_DAYS:
            reason = f"חדש {age} ימים בלי הגשה"
        elif status in ("נדחה", "דילגתי") and age >= STALE_CLOSED_DAYS:
            reason = f"{status}, בן {age} ימים"
        if reason:
            vals = [c.value for c in row[:13]]
            if row[7].hyperlink:  # legacy rows: real URL lives on the hyperlink
                vals[7] = row[7].hyperlink.target
            stale.append((row[0].row, vals, reason))

    moved = [f"#{v[0]} {v[2]} ({r})" for _, v, r in stale]
    # second pass: archive rows past their week get deleted for good (tombstone
    # in Deleted keeps them out of future scans). Applied/interview rows stay.
    purge = []  # (archive_row_idx, tombstone values, label)
    for row in ar.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        status = str(row[9].value or "")
        if status in ("הוגש", "ראיון") or row[12].value:
            continue
        arc = parse_date(row[13].value) if len(row) > 13 else None
        if arc is None or (today - arc).days < PURGE_ARCHIVE_DAYS:
            continue
        link = row[7].hyperlink.target if row[7].hyperlink else row[7].value
        purge.append((row[0].row,
                      [today.isoformat(), row[2].value, row[3].value, link, row[8].value],
                      f"#{row[0].value} {row[2].value} (נמחק - {(today - arc).days} ימים בארכיון)"))
    moved += [p[2] for p in purge]
    if not dry_run:
        for _, vals, reason in stale:
            vals = list(vals) + [None] * (13 - len(vals))
            ar.append(vals + [today.isoformat(), reason])
        for idx, _, _ in sorted(stale, key=lambda s: s[0], reverse=True):
            ws.delete_rows(idx)
        if purge:
            if DELETED_SHEET not in wb.sheetnames:
                dl = wb.create_sheet(DELETED_SHEET)
                dl.sheet_view.rightToLeft = True
                for col, head in enumerate(("תאריך מחיקה", "חברה", "משרה", "קישור", "Job ID"), start=1):
                    dl.cell(row=1, column=col, value=head)
            dl = wb[DELETED_SHEET]
            for _, vals, _ in purge:
                dl.append(vals)
            for idx, _, _ in sorted(purge, key=lambda p: p[0], reverse=True):
                ar.delete_rows(idx)
        if stale or purge or schema_changed:
            wb.save(TRACKER)
    return len(stale) + len(purge), moved


def main():
    dry = "--dry-run" in sys.argv
    n, moved = run(dry_run=dry)
    tag = "would archive" if dry else "archived"
    print(f"{tag}: {n}")
    for m in moved:
        print(" ", m)


if __name__ == "__main__":
    main()
