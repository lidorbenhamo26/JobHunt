# -*- coding: utf-8 -*-
"""Rebuild dashboard.html from jobs.xlsx.

Usage: python make_dashboard.py
RTL Hebrew dashboard, Stitch-inspired design (Material-ish, self-contained CSS,
no CDN so file:// works offline): big score badge, corner ribbons, AI-insight
box, sticky toolbar (search / sort / filters), "הגש הכל" batch pipeline,
auto-archive pass on every rebuild, live task console. Light theme only.
Icons are inline SVG (Material-style) - no decorative emoji, per user
preference.
"""
import html
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

import archive_stale

ROOT = Path(__file__).resolve().parent.parent
TRACKER = ROOT / "jobs.xlsx"
OUT = ROOT / "dashboard.html"
STALE_NEW_DAYS = archive_stale.STALE_NEW_DAYS
FOLLOWUP_DAYS = 10  # "הוגש" with no reply for this long -> nudge

# Inline Material-style SVG icons (fill=currentColor, scale with font)
def _svg(path, vb="0 0 24 24"):
    return (f'<svg class="ic" viewBox="{vb}" xmlns="http://www.w3.org/2000/svg" '
            f'fill="currentColor" aria-hidden="true"><path d="{path}"/></svg>')


IC_PIN = _svg("M12 2C8.13 2 5 5.13 5 9c0 5.25 7 13 7 13s7-7.75 7-13c0-3.87-3.13-7-7-7zm0 9.5a2.5 2.5 0 1 1 0-5 2.5 2.5 0 0 1 0 5z")
IC_CAL = _svg("M19 4h-1V2h-2v2H8V2H6v2H5c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h14c1.1 0 2-.9 2-2V6c0-1.1-.9-2-2-2zm0 16H5V10h14v10zM5 8V6h14v2H5z")
IC_LINK = _svg("M14 3h7v7h-2V6.41l-9.29 9.3-1.42-1.42L17.59 5H14V3zM5 5h6v2H7v10h10v-4h2v6H5V5z")
IC_TRASH = _svg("M9 3h6l1 2h4v2H4V5h4l1-2zm-2 6h10l-.9 12H8L7 9zm3 2v8h1v-8h-1zm3 0v8h1v-8h-1z")
IC_SPARK = _svg("M12 2l1.9 5.6L19.5 9.5l-5.6 1.9L12 17l-1.9-5.6L4.5 9.5l5.6-1.9L12 2zm7.5 11l.9 2.6 2.6.9-2.6.9-.9 2.6-.9-2.6-2.6-.9 2.6-.9.9-2.6z")
IC_GRID = _svg("M3 3h8v8H3V3zm10 0h8v5h-8V3zM3 13h8v8H3v-8zm10-3h8v11h-8V10z")
IC_BOX = _svg("M3 4h18v4H3V4zm2 6h14v10H5V10zm4 3h6v2H9v-2z")
IC_USER = _svg("M12 12a5 5 0 1 0-5-5 5 5 0 0 0 5 5zm0 2c-4.42 0-8 2.24-8 5v1h16v-1c0-2.76-3.58-5-8-5z")
IC_BOLT = _svg("M7 2v11h3v9l7-12h-4l4-8H7z")
IC_DOC = _svg("M14 2H6c-1.1 0-2 .9-2 2v16c0 1.1.9 2 2 2h12c1.1 0 2-.9 2-2V8l-6-6zm2 16H8v-2h8v2zm0-4H8v-2h8v2zm-3-5V3.5L18.5 9H13z")
IC_SEARCH = _svg("M15.5 14h-.79l-.28-.27A6.47 6.47 0 0 0 16 9.5 6.5 6.5 0 1 0 9.5 16c1.61 0 3.09-.59 4.23-1.57l.27.28v.79l5 4.99L20.49 19l-4.99-5zm-6 0C7.01 14 5 11.99 5 9.5S7.01 5 9.5 5 14 7.01 14 9.5 11.99 14 9.5 14z")


# On-demand scan controls. Shared verbatim by the dashboard and the sub-pages,
# so it may only depend on BASE + toast() (both exist on every page).
SCAN_JS = """
let scanWas = null;
async function runScan(mode) {
  const msg = mode === "cheap"
    ? "להריץ סריקה מהירה עכשיו?\\n\\nNVIDIA / Intel / Amazon / לינקדאין (חיפה והצפון) / ג'וב מאסטר,\\nרק משרות מ-72 השעות האחרונות. כמה דקות, רץ ברקע."
    : "להריץ סריקה מלאה עכשיו?\\n\\nכל המקורות שנגישים ב-HTTP, חלון של 7 ימים - לוקח יותר זמן.\\n\\nשים לב: מקורות שדורשים דפדפן (מיקרוסופט, גוגל, אפל, צ'ק פוינט, דרושים, אולג'ובס, וואטסאפ) לא רצים מכאן, ולכן התאריך של הסריקה המלאה המתוזמנת לא מתעדכן - היא עדיין תרוץ כרגיל.";
  if (!confirm(msg)) return;
  let s;
  try {
    s = await (await fetch(BASE + "/scan?mode=" + mode, { signal: AbortSignal.timeout(5000) })).json();
  } catch (e) {
    return toast("השרת המקומי כבוי - דאבל-קליק על start-cv-server.bat בתיקיית JobHunt ונסה שוב");
  }
  toast(s.detail ? s.detail
        : (mode === "cheap" ? "סריקה מהירה יצאה לדרך - ההתקדמות בקונסול המשימות למטה"
                            : "סריקה מלאה יצאה לדרך - ההתקדמות בקונסול המשימות למטה"));
  scanPoll();
}
async function scanPoll() {
  let s;
  try { s = await (await fetch(BASE + "/scan-status", { signal: AbortSignal.timeout(2500) })).json(); }
  catch (e) { return; }
  const run = s.status === "running";
  for (const id of ["scan-cheap", "scan-full"]) {
    const b = document.getElementById(id);
    if (b) b.disabled = run;
  }
  const el = document.getElementById("scanst");
  if (el) {
    el.className = "scanst" + (run ? " run" : (s.status === "error" ? " err" : ""));
    if (run) {
      const mins = s.started ? Math.max(0, Math.round((Date.now() / 1000 - s.started) / 60)) : 0;
      el.textContent = (s.mode === "full" ? "סריקה מלאה רצה" : "סריקה מהירה רצה") + " · " + mins + " דק'";
    } else if (s.status === "done") {
      el.textContent = "הסריקה הסתיימה";
    } else if (s.status === "error") {
      el.textContent = "הסריקה נכשלה: " + String(s.detail || "").slice(0, 90);
    } else {
      el.textContent = "";
    }
  }
  if (run) { scanWas = s.status; return setTimeout(scanPoll, 5000); }
  if (scanWas === "running" && s.status === "done") {
    toast("הסריקה הסתיימה - מרענן את הלוח");
    setTimeout(() => location.reload(), 1800);
  }
  scanWas = s.status;
}
"""


def parse_d(s):
    try:
        return date.fromisoformat(str(s)[:10])
    except (ValueError, TypeError):
        return None


def age_text(days):
    if days <= 0:
        return "היום"
    if days == 1:
        return "אתמול"
    return f"לפני {days} ימים"


def load_submit_sessions():
    """Job ids with a saved Codex submit session (resumable via המשך)."""
    try:
        raw = (ROOT / "output" / ".submit_sessions.json").read_text(encoding="utf-8-sig")
        return {int(k) for k in json.loads(raw)}
    except Exception:  # noqa: BLE001
        return set()


STATUS_COLORS = {
    "חדש": ("#e8f1fb", "#2563a8"),
    "הוגש": ("#e6f6ea", "#1a7f37"),
    "ראיון": ("#fff3e0", "#b45309"),
    "נדחה": ("#fdecec", "#b42318"),
    "דילגתי": ("#eceef1", "#57606a"),
}


def load_jobs():
    wb = load_workbook(TRACKER)
    ws = wb["Jobs"]
    jobs = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        jobs.append({
            "id": row[0].value, "date": str(row[1].value or "")[:10],
            "company": row[2].value or "", "title": row[3].value or "",
            "location": row[4].value or "", "score": row[5].value or 0,
            "reason": row[6].value or "",
            "link": row[7].hyperlink.target if row[7].hyperlink else (row[7].value or ""),
            "job_id": row[8].value or "", "status": row[9].value or "חדש",
            "cv": row[10].value or "",
            "submitted": str(row[12].value)[:10] if len(row) > 12 and row[12].value else "",
        })
    archived = 0
    if "Archive" in wb.sheetnames:
        archived = sum(1 for r in wb["Archive"].iter_rows(min_row=2, values_only=True)
                       if r[0] is not None)
    return sorted(jobs, key=lambda j: (-(j["score"] or 0), j["id"])), archived


def build(jobs, archived=0):
    resumable = load_submit_sessions()
    today = date.today()
    cards = []
    for j in jobs:
        bg, fg = STATUS_COLORS.get(j["status"], ("#eceef1", "#57606a"))
        d = parse_d(j["date"])
        age = (today - d).days if d else None
        date_meta = f'{IC_CAL} {html.escape(j["date"])}' + (f' · {age_text(age)}' if age is not None else "")
        ribbon = ""
        if j["status"] == "חדש" and age is not None:
            left = STALE_NEW_DAYS - age
            if left <= 0:
                ribbon = '<span class="rib rib-y">ייכנס לארכיון בסריקה הבאה</span>'
            elif left <= 7:
                ribbon = f'<span class="rib rib-y">ארכוב בעוד {left} ימים - להגיש או לדלג</span>'
        sub_meta = ""
        sub_d = parse_d(j["submitted"])
        if sub_d and j["status"] in ("הוגש", "ראיון"):
            sub_age = (today - sub_d).days
            sub_meta = f'<span class="dot">•</span><span class="meta">הוגש {html.escape(j["submitted"])}</span>'
            if j["status"] == "הוגש" and sub_age >= FOLLOWUP_DAYS:
                ribbon = f'<span class="rib rib-p">{sub_age} ימים בלי מענה - שווה פולו-אפ</span>'
        cv_html = ""
        if j["cv"]:
            # Relative path - opens the PDF straight off disk from file://.
            cv_html = f'<a class="btn cvready" href="output/{html.escape(str(j["cv"]), quote=True)}" target="_blank" title="פתיחת ה-CV המוכן (PDF)">{IC_DOC} CV מוכן</a>'
        status_opts = "".join(
            f'<option value="{s}"{" selected" if j["status"] == s else ""}>{s}</option>'
            for s in STATUS_COLORS
        )
        submit_btn = ""
        continue_btn = ""
        if j["cv"] and j["status"] not in ("הוגש", "ראיון"):
            # single button; its label/behavior follow the sidebar mode switch
            submit_btn = (
                f'<button class="btn primary sub-go" onclick="submitJob({j["id"]}, this)" '
                f'title="Codex ממלא ועוצר לפני השליחה - אתה בודק בכרום ולוחץ Submit בעצמך">הגש מועמדות</button>'
            )
            if j["id"] in resumable:
                continue_btn = f'<button class="btn outline" onclick="continueSubmit({j["id"]}, this)" title="ממשיך את ריצת ה-Codex הקודמת מאיפה שנעצרה">המשך</button>'
        # CV exists -> regenerate is a secondary action (outline), not the loud navy CTA
        if j["cv"]:
            gen_btn = (f'<button class="btn outline" onclick="genCV({j["id"]}, this)" '
                       f'title="יצירה מחדש של ה-CV (ידרוס את הקיים)">הפק CV מחדש</button>')
        else:
            gen_btn = f'<button class="btn navy" onclick="genCV({j["id"]}, this)">הפק קורות חיים</button>'
        score_cls = min(int(j["score"] or 0), 10)
        # search haystack: content fields only, so typing "cv" / "הגש" won't
        # match every card via its button labels
        blob = " ".join(str(j[k]) for k in ("company", "title", "reason", "location"))
        blob = f'{blob} #{j["id"]}'.lower()
        cards.append(f"""
    <div class="card{' has-rib' if ribbon else ''}" id="job-{j['id']}" data-status="{html.escape(str(j['status']))}"
         data-score="{j['score'] or 0}" data-date="{html.escape(j['date'])}" data-search="{html.escape(blob, quote=True)}">
      {ribbon}
      <div class="score s{score_cls}"><span class="sn">{j['score']}</span><span class="sl">התאמה</span></div>
      <div class="body">
        <div class="top">
          <input type="checkbox" class="selbox" data-id="{j['id']}" onchange="selChanged()" title="בחירה להגשה/CV מרובים" aria-label="בחירה להגשה או CV מרובים"{' disabled' if j['status'] in ('הוגש', 'ראיון') else ''}>
          <span class="company">{html.escape(str(j['company']))}</span>
          <select class="badge badge-sel" style="background:{bg};color:{fg}" onchange="setStatus({j['id']}, this.value, this)" title="שינוי סטטוס" aria-label="סטטוס המשרה">{status_opts}</select>
          <span class="meta">{IC_PIN} {html.escape(str(j['location']))}</span>
          <span class="dot">•</span>
          <span class="meta">{date_meta}</span>
          {sub_meta}
          <span class="dot">•</span>
          <span class="meta mut">#{j['id']}</span>
          <span class="toplinks">{cv_html}<a class="btn outline iconb" href="{html.escape(str(j['link']), quote=True)}" target="_blank" title="פתיחת המשרה במקור" aria-label="פתיחת המשרה במקור">{IC_LINK}</a></span>
        </div>
        <div class="title">{html.escape(str(j['title']))}</div>
        <div class="insight" onclick="this.classList.toggle('open')" title="קליק להרחבה"><span class="ri">{IC_SPARK}</span><span class="reason">{html.escape(str(j['reason']))}</span></div>
        <div class="actions">
          {gen_btn}
          {submit_btn}
          {continue_btn}
          <button class="btn del iconb" onclick="delJob({j['id']}, this)" title="מחיקת המשרה מהטבלה" aria-label="מחיקת המשרה מהטבלה">{IC_TRASH}</button>
        </div>
      </div>
    </div>""")

    counts = {s: sum(1 for j in jobs if j["status"] == s) for s in STATUS_COLORS}
    new_count = counts["חדש"]
    applied = counts["הוגש"] + counts["ראיון"]
    top = sum(1 for j in jobs if (j["score"] or 0) >= 8)

    # focus: top 3 to apply today - best score, freshest first on ties
    cand = [j for j in jobs if j["status"] == "חדש"]
    cand.sort(key=lambda j: j["date"], reverse=True)
    cand.sort(key=lambda j: j["score"] or 0, reverse=True)
    focus_html = ""
    if cand:
        minis = "".join(
            f'<a class="fcard" href="#job-{j["id"]}">'
            f'<span class="fscore s{min(int(j["score"] or 0), 10)}">{j["score"]}</span>'
            f'<span class="ftxt"><span class="fc">{html.escape(str(j["company"]))}</span>'
            f'<span class="ft">{html.escape(str(j["title"])[:44])}</span></span></a>'
            for j in cand[:3])
        focus_html = (f'<div class="focus"><div class="focus-t">מומלץ להיום</div>'
                      f'<div class="fgrid">{minis}</div></div>')
    archive_chip = f'<a class="chip chip-link" href="archive.html">בארכיון <b>{archived}</b></a>' if archived else ""

    filters = "".join(
        f'<button class="f" data-f="{s}" onclick="filt(this)">{s} <span class="cnt">{counts[s]}</span></button>'
        for s in ("חדש", "הוגש", "ראיון"))

    empty = ('<div class="empty">הרשימה נקייה.<br>הסריקה הבאה תביא משרות לפי '
             'הקריטריונים החדשים (0-3 שנים, כולל firmware/low-level).<br>'
             '<span class="mut">סריקה ידנית: תגיד ל-Claude "תריץ סריקת משרות עכשיו"</span></div>')

    return f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>JobHunt - לידור</title>
<style>
  :root {{
    --bg:#f8fafc; --card:#ffffff; --line:#e2e8f0; --line2:#cbd5e1; --ink:#091426;
    --mut:#45474c; --soft:#f0edef; --softer:#f5f3f4;
    --navy:#091426; --navy-h:#1e293b; --navy-ink:#ffffff;
    --pri:#712ae2; --pri-h:#5e1fca; --pri-ink:#ffffff;
    --mint:#6ffbbe; --mint-dim:#4edea3; --mint-ink:#002113;
    --gray-badge:#eae7e9; --gray-badge-ink:#091426;
    --ok:#1a7f37; --okbg:#e6f6ea; --err:#ba1a1a; --errbg:#ffdad6;
    --rib-y-bg:#fef08a; --rib-y-ink:#854d0e; --rib-p-bg:#eaddff; --rib-p-ink:#25005a;
    --insight-bg:#f8fafc; --insight-line:#cbd5e1;
    --shadow:0 4px 12px rgba(30,41,59,.05); --shadow-hover:0 6px 18px rgba(30,41,59,.12);
  }}
  /* light mode only - per user preference (2026-08-10) */
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  html {{ color-scheme: light; }}
  body {{ font-family:"Work Sans","Segoe UI",Arial,sans-serif; background:var(--bg); color:var(--ink); }}
  .ic {{ width:1em; height:1em; vertical-align:-0.13em; flex-shrink:0; }}
  :focus-visible {{ outline:2px solid var(--pri); outline-offset:2px; }}
  /* ---- layout: fixed sidebar (right, RTL) + main ---- */
  aside {{ position:fixed; top:0; right:0; bottom:0; width:250px; z-index:50;
           background:var(--card); border-left:1px solid var(--line);
           display:flex; flex-direction:column; padding:18px 14px;
           box-shadow:0 0 12px rgba(30,41,59,.04); }}
  .brand {{ font-size:17px; font-weight:800; letter-spacing:-.02em; padding:2px 10px 14px; }}
  .userbox {{ display:flex; flex-direction:column; gap:2px; background:var(--softer);
              border:1px solid var(--line); border-radius:12px; padding:10px 12px; margin-bottom:16px; }}
  .uname {{ font-weight:700; font-size:14px; }}
  .urole {{ color:var(--mut); font-size:12px; }}
  nav.side {{ display:flex; flex-direction:column; gap:4px; }}
  .navi {{ display:flex; align-items:center; gap:10px; padding:10px 12px; border-radius:12px;
           text-decoration:none; color:var(--mut); font-size:13.5px; font-weight:600;
           transition:background .12s; }}
  .navi:hover {{ background:var(--softer); }}
  .navi.act {{ background:var(--pri); color:var(--pri-ink); }}
  .navi .cnt {{ margin-right:auto; background:var(--soft); border-radius:9999px; padding:1px 8px;
                font-size:12px; color:var(--mut); }}
  .navi.act .cnt {{ background:rgba(255,255,255,.2); color:var(--pri-ink); }}
  .sidegrow {{ flex:1; }}
  .sideact {{ display:flex; flex-direction:column; gap:8px; }}
  .bigbtn {{ border:none; border-radius:10px; padding:11px 16px; font-size:13.5px; font-weight:700;
             cursor:pointer; font-family:inherit; background:var(--pri); color:var(--pri-ink);
             box-shadow:var(--shadow); transition:transform .08s; width:100%; }}
  .bigbtn:hover {{ background:var(--pri-h); }}
  .bigbtn:active {{ transform:scale(.98); }}
  .bigbtn.sec {{ background:var(--card); color:var(--ink); border:1px solid var(--line2); box-shadow:none; }}
  .bigbtn.sec:hover {{ background:var(--softer); }}
  .bigbtn:disabled {{ opacity:.55; cursor:default; }}
  .bigbtn.dim {{ opacity:.55; }}
  .swrow {{ display:flex; align-items:center; gap:8px; cursor:pointer; padding:0 2px 2px; user-select:none; }}
  .swrow input {{ display:none; }}
  .sw {{ width:34px; height:19px; border-radius:10px; background:var(--line2); position:relative;
         transition:background .15s; flex:none; }}
  .sw::after {{ content:""; position:absolute; top:2px; right:2px; width:15px; height:15px; border-radius:50%;
                background:#fff; transition:transform .15s; box-shadow:0 1px 2px rgba(0,0,0,.25); }}
  .swrow input:checked ~ .sw {{ background:var(--pri); }}
  .swrow input:checked ~ .sw::after {{ transform:translateX(-15px); }}
  .swtxt {{ font-size:12.5px; font-weight:600; color:var(--mut); display:flex; align-items:center; gap:6px; }}
  .swrow input:checked ~ .swtxt {{ color:var(--pri); }}
  /* on-demand job scan (same routine the scheduled task runs) */
  .scangrp {{ border-top:1px solid var(--line); margin-top:4px; padding-top:10px; }}
  .scanh {{ display:flex; align-items:center; gap:6px; font-size:12px; font-weight:700;
            color:var(--mut); padding:0 2px 8px; }}
  .scanrow {{ display:flex; gap:8px; }}
  .scanrow .bigbtn {{ flex:1; padding:10px 6px; font-size:12.5px; }}
  .bigbtn.scan {{ background:var(--navy); color:var(--navy-ink); }}
  .bigbtn.scan:hover {{ background:var(--navy-h); }}
  .bigbtn.sec.scan {{ background:var(--card); color:var(--ink); border:1px solid var(--line2); }}
  .bigbtn.sec.scan:hover {{ background:var(--softer); }}
  .scanst {{ font-size:11.5px; color:var(--mut); line-height:1.6; padding:7px 2px 0; }}
  .scanst.run {{ color:var(--pri); font-weight:600; }}
  .scanst.err {{ color:var(--err); }}
  .review-box {{ margin-top:10px; border:1px solid var(--line2); border-radius:12px;
                 background:var(--softer); padding:10px 12px; }}
  .review-box .rvh {{ font-size:12px; font-weight:700; margin-bottom:8px; }}
  .review-box pre {{ margin:0 0 8px; max-height:260px; overflow:auto; white-space:pre-wrap;
                     font-size:11.5px; line-height:1.5; background:var(--card);
                     border:1px solid var(--line); border-radius:8px; padding:8px 10px; }}
  .review-box .rvimg {{ display:block; max-width:100%; max-height:320px; object-fit:contain;
                        border:1px solid var(--line); border-radius:8px; margin-bottom:8px;
                        cursor:zoom-in; }}
  .rvb {{ display:flex; gap:8px; }}
  .ans-box .ans-q {{ font-size:12.5px; font-weight:600; margin-bottom:8px; background:var(--card);
                     border:1px solid var(--line); border-radius:8px; padding:8px 10px; }}
  .ans-box textarea {{ width:100%; box-sizing:border-box; resize:vertical; min-height:44px;
                       font:inherit; font-size:12.5px; background:var(--card); color:inherit;
                       border:1px solid var(--line); border-radius:8px; padding:8px 10px;
                       margin-bottom:8px; }}
  .sidefoot {{ color:var(--mut); font-size:12px; line-height:1.7; padding:12px 10px 0; }}
  main {{ margin-right:250px; padding:22px 26px 120px; }}
  .wrap {{ max-width:960px; margin:0 auto; }}
  h1 {{ font-size:24px; font-weight:700; letter-spacing:-.02em; }}
  .chipsrow {{ display:flex; gap:8px; margin:12px 0 16px; flex-wrap:wrap; align-items:center; }}
  .chip {{ background:var(--card); border:1px solid var(--line); border-radius:9999px;
           padding:5px 14px; font-size:12.5px; color:var(--mut); box-shadow:var(--shadow); }}
  .chip b {{ color:var(--ink); }}
  .chip.mintc {{ background:var(--mint-dim); color:var(--mint-ink); }}
  .chip.mintc b {{ color:var(--mint-ink); }}
  .chip.pric {{ background:var(--rib-p-bg); color:var(--rib-p-ink); }}
  .chip.pric b {{ color:var(--rib-p-ink); }}
  .chip-link {{ text-decoration:none; }}
  .chip-link:hover {{ background:var(--softer); }}
  /* search / sort / filters row - sticky inside main */
  .listbar {{ position:sticky; top:0; z-index:30; background:var(--bg); display:flex; gap:8px;
              flex-wrap:wrap; align-items:center; padding:10px 0 12px; margin-bottom:6px;
              border-bottom:1px solid var(--line); }}
  #q {{ border:1px solid var(--line); background:var(--card); color:var(--ink);
        border-radius:9999px; padding:9px 14px; font-size:13px; font-family:inherit; width:210px; }}
  #sortsel {{ border:1px solid var(--line); background:var(--card); color:var(--ink);
              border-radius:9999px; padding:9px 12px; font-size:12.5px; font-family:inherit; cursor:pointer; }}
  .f {{ border:1px solid var(--line); background:var(--card); color:var(--ink); border-radius:9999px;
        padding:8px 14px; font-size:12.5px; cursor:pointer; font-family:inherit; box-shadow:var(--shadow);
        transition:transform .08s; }}
  .f:active {{ transform:scale(.95); }}
  .f .cnt {{ background:var(--soft); border-radius:9999px; padding:1px 7px; font-size:12px; color:var(--mut); }}
  .f.on {{ background:var(--navy); color:var(--navy-ink); border-color:var(--navy); }}
  .f.on .cnt {{ background:rgba(255,255,255,.18); color:var(--navy-ink); }}
  /* ---- batch progress ---- */
  #batch {{ display:none; background:var(--card); border:1px solid var(--line); border-radius:12px;
            padding:12px 16px; margin-bottom:16px; font-size:13px; box-shadow:var(--shadow); }}
  #batch .bph {{ font-weight:700; }}
  #batch .bstop {{ float:left; border:1px solid var(--err); color:var(--err); background:none;
                   border-radius:9999px; padding:3px 12px; cursor:pointer; font-family:inherit; font-size:12px; }}
  #bprog {{ height:6px; background:var(--soft); border-radius:4px; margin-top:8px; overflow:hidden; }}
  #bfill {{ height:100%; width:0%; background:var(--pri); transition:width .4s; }}
  /* ---- focus section ---- */
  .focus {{ background:var(--card); border:1px solid var(--line); border-radius:12px;
            padding:14px 16px; margin-bottom:16px; box-shadow:var(--shadow); }}
  .focus-t {{ font-weight:600; font-size:15px; margin-bottom:10px; }}
  .fgrid {{ display:grid; grid-template-columns:repeat(3, 1fr); gap:10px; }}
  .fcard {{ display:flex; gap:10px; align-items:flex-start; border:1px solid var(--line);
            border-radius:10px; padding:9px 11px; text-decoration:none; color:var(--ink);
            transition:background .15s; }}
  .fcard:hover {{ background:var(--softer); }}
  .fscore {{ font-weight:700; font-size:14px; border-radius:8px; padding:3px 9px;
             background:var(--gray-badge); color:var(--gray-badge-ink); }}
  .fscore.s10,.fscore.s9 {{ background:var(--mint); color:var(--mint-ink); }}
  .fscore.s8 {{ background:var(--mint-dim); color:var(--mint-ink); }}
  .ftxt {{ display:flex; flex-direction:column; min-width:0; }}
  .fc {{ font-weight:600; font-size:13.5px; }}
  .ft {{ color:var(--mut); font-size:12px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
  /* ---- cards ---- */
  .card {{ position:relative; display:flex; gap:14px; background:var(--card); border:1px solid var(--line);
           border-radius:12px; padding:16px; margin-bottom:12px; box-shadow:var(--shadow);
           scroll-margin-top:160px; transition:box-shadow .15s, transform .15s; }}
  .card:hover {{ box-shadow:var(--shadow-hover); transform:translateY(-1px); }}
  .card:target {{ outline:2px solid var(--pri); }}
  .card.has-rib {{ padding-top:38px; }}
  .rib {{ position:absolute; top:8px; left:8px; font-size:12px; font-weight:700; letter-spacing:.02em;
          border-radius:9999px; padding:3px 12px; box-shadow:var(--shadow); }}
  .rib-y {{ background:var(--rib-y-bg); color:var(--rib-y-ink); }}
  .rib-p {{ background:var(--rib-p-bg); color:var(--rib-p-ink); }}
  .score {{ display:flex; flex-direction:column; align-items:center; justify-content:center;
            min-width:72px; height:72px; border-radius:12px; gap:1px;
            background:var(--gray-badge); color:var(--gray-badge-ink); }}
  .score .sn {{ font-size:24px; font-weight:800; letter-spacing:-.02em; }}
  .score .sl {{ font-size:11px; font-weight:700; letter-spacing:.06em; opacity:.8; }}
  .s10,.s9 {{ background:var(--mint); color:var(--mint-ink); }}
  .s8 {{ background:var(--mint-dim); color:var(--mint-ink); }}
  .s7 {{ background:var(--gray-badge); color:var(--gray-badge-ink); }}
  .body {{ flex:1; min-width:0; }}
  .top {{ display:flex; gap:8px; align-items:center; flex-wrap:wrap; }}
  /* view-links (CV / מקור) pinned to the card's top-left corner; RTL: margin-right:auto pushes left */
  .toplinks {{ display:inline-flex; gap:6px; align-items:center; margin-right:auto; }}
  .toplinks .btn {{ padding:5px 12px; font-size:12px; border-radius:9999px; }}
  .toplinks .iconb {{ padding:5px 9px; font-size:13px; }}
  .company {{ font-weight:700; font-size:14.5px; }}
  .badge {{ font-size:12px; border-radius:9999px; padding:3px 11px; font-weight:600; }}
  .badge-sel {{ border:none; cursor:pointer; font-family:inherit; }}
  .selbox {{ width:17px; height:17px; accent-color:var(--pri); cursor:pointer; flex-shrink:0; }}
  .selbox:disabled {{ opacity:.35; cursor:default; }}
  .meta {{ color:var(--mut); font-size:12px; display:inline-flex; align-items:center; gap:4px; }}
  .dot {{ color:var(--mut); opacity:.5; font-size:11px; }}
  .mut {{ opacity:.75; }}
  .title {{ font-size:16px; font-weight:600; margin:6px 0 6px; }}
  .insight {{ display:flex; gap:8px; align-items:flex-start; background:var(--insight-bg);
              border:1px solid var(--insight-line); border-radius:10px; padding:8px 11px;
              margin-bottom:10px; cursor:pointer; }}
  .insight .ri {{ color:var(--pri); font-size:13px; line-height:1.4; }}
  .insight::after {{ content:"▾"; color:var(--mut); font-size:12px; margin-right:auto;
                     align-self:center; transition:transform .15s; }}
  .insight.open::after {{ transform:rotate(180deg); }}
  .reason {{ color:var(--ink); font-size:12.8px; opacity:.85;
             display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; overflow:hidden; }}
  .insight.open .reason {{ -webkit-line-clamp:unset; }}
  .actions {{ display:flex; gap:8px; flex-wrap:wrap; align-items:center; }}
  .btn {{ border:none; border-radius:9px; padding:8px 15px; font-size:13px; font-weight:600; cursor:pointer;
          text-decoration:none; display:inline-flex; align-items:center; gap:5px; font-family:inherit;
          box-shadow:var(--shadow); transition:transform .08s; }}
  .btn:active {{ transform:scale(.96); }}
  .btn.primary {{ background:var(--pri); color:var(--pri-ink); font-weight:700; }}
  .btn.primary:hover {{ background:var(--pri-h); }}
  .btn.navy {{ background:var(--navy); color:var(--navy-ink); }}
  .btn.navy:hover {{ background:var(--navy-h); }}
  .btn.outline {{ background:var(--card); color:var(--ink); border:1px solid var(--line2); box-shadow:none; }}
  .btn.outline:hover {{ background:var(--softer); }}
  .btn.cvready {{ background:var(--mint); color:var(--mint-ink); font-weight:700;
                  box-shadow:0 1px 4px rgba(0,33,19,.18); }}
  .btn.cvready:hover {{ background:var(--mint-dim); }}
  .btn.del {{ background:none; color:var(--err); margin-right:auto; box-shadow:none; }}
  .btn.del:hover {{ background:var(--errbg); }}
  .btn.iconb {{ padding:8px 10px; font-size:15px; }}
  .btn:disabled {{ opacity:.6; cursor:default; }}
  .empty {{ text-align:center; color:var(--mut); padding:60px 0; font-size:15px; line-height:2; }}
  .toast {{ position:fixed; bottom:76px; right:50%; transform:translateX(50%); background:var(--navy);
            color:var(--navy-ink); padding:10px 22px; border-radius:10px; font-size:14px; opacity:0;
            transition:opacity .25s; pointer-events:none; z-index:70; max-width:90vw; box-shadow:var(--shadow-hover); }}
  .toast.show {{ opacity:1; }}
  /* ---- task console ---- */
  #cbar {{ position:fixed; bottom:0; left:0; right:250px; z-index:60; background:#0f172a; color:#cbd5e1;
           box-shadow:0 -4px 18px rgba(2,6,23,.35); font-size:12.5px;
           font-family:Consolas,"JetBrains Mono","Courier New",monospace; }}
  .cb-head {{ display:flex; gap:10px; align-items:center; padding:8px 16px; cursor:pointer; user-select:none; }}
  .cb-title {{ font-weight:700; color:#e2e8f0; white-space:nowrap; }}
  #cb-chips {{ display:flex; gap:8px; flex-wrap:wrap; flex:1; }}
  .cb-chip {{ display:inline-flex; gap:6px; align-items:center; background:#1e293b; border:1px solid #334155;
              border-radius:6px; padding:3px 11px; white-space:nowrap; cursor:pointer; }}
  .cb-chip.running {{ border-color:#38bdf8; color:#bae6fd; }}
  .cb-chip.ok {{ border-color:#4ade80; color:#bbf7d0; }}
  .cb-chip.warn {{ border-color:#fbbf24; color:#fde68a; }}
  .cb-chip.err {{ border-color:#f87171; color:#fecaca; }}
  .cb-idle {{ color:#94a3b8; }}
  .cb-arrow {{ color:#64748b; }}
  .cb-spin {{ display:inline-block; animation:cbrot 1.1s linear infinite; }}
  @keyframes cbrot {{ to {{ transform:rotate(360deg); }} }}
  #cb-panel {{ display:none; border-top:1px solid #1e293b; padding:10px 16px 12px; }}
  #cb-tabs {{ display:flex; gap:8px; flex-wrap:wrap; margin-bottom:8px; }}
  .cb-tab {{ background:#1e293b; border:1px solid #334155; color:#cbd5e1; border-radius:6px; padding:4px 10px;
             cursor:pointer; font-family:inherit; font-size:12px; }}
  .cb-tab.sel {{ background:#334155; border-color:#38bdf8; color:#fff; }}
  #cb-detail {{ display:none; background:#450a0a; border:1px solid #b91c1c; color:#fecaca; border-radius:8px;
                padding:7px 12px; margin-bottom:8px; white-space:pre-wrap; }}
  #cb-detail.warn {{ background:#451a03; border-color:#b45309; color:#fde68a; }}
  #cb-log {{ background:#020617; color:#94a3b8; border-radius:8px; padding:10px 12px; height:30vh; overflow:auto;
             font-size:12px; white-space:pre-wrap; word-break:break-word; margin:0; }}
  .cb-actions {{ display:flex; gap:14px; align-items:center; margin-top:8px; color:#94a3b8; font-size:12px; }}
  .cb-btn {{ background:#1e293b; border:1px solid #334155; color:#cbd5e1; border-radius:6px; padding:4px 12px;
             cursor:pointer; font-family:inherit; font-size:12px; }}
  .cb-btn:hover {{ background:#334155; }}
  @media (prefers-reduced-motion:reduce) {{
    *, *::before, *::after {{ transition:none !important; animation:none !important; }}
  }}
  @media (max-width:900px) {{
    aside {{ position:static; width:auto; border-left:none; border-bottom:1px solid var(--line);
             flex-direction:row; flex-wrap:wrap; align-items:center; gap:8px; padding:12px 16px; }}
    .brand {{ padding:0 0 0 6px; }}
    .userbox {{ flex-direction:row; gap:8px; align-items:baseline; margin:0; padding:6px 12px; }}
    nav.side {{ flex-direction:row; flex-wrap:wrap; }}
    .navi {{ padding:8px 12px; }}
    .sidegrow {{ display:none; }}
    .sideact {{ flex-direction:row; }}
    .bigbtn {{ width:auto; }}
    .sidefoot {{ display:none; }}
    main {{ margin-right:0; padding:16px 14px 120px; }}
    #cbar {{ right:0; }}
  }}
  @media (max-width:640px) {{
    .card {{ flex-direction:column; }}
    .card.has-rib {{ padding-top:40px; }}
    .score {{ flex-direction:row; min-width:0; width:fit-content; height:auto; padding:5px 12px; gap:6px; }}
    .score .sn {{ font-size:16px; }}
    .fgrid {{ grid-template-columns:1fr; }}
    #q {{ width:100%; }}
  }}
  /* touch devices (phone via LAN/Tailscale): 44px-ish targets, desktop untouched.
     Last so it wins over the width breakpoints above. */
  @media (pointer:coarse) {{
    .btn {{ padding:13px 16px; }}
    .btn.iconb {{ padding:13px 14px; }}
    .toplinks .btn {{ padding:10px 14px; }}
    .toplinks .iconb {{ padding:10px 12px; }}
    .selbox {{ width:22px; height:22px; }}
    .badge {{ font-size:13px; padding:8px 13px; }}
    .f {{ padding:13px 16px; }}
    #q {{ padding:12px 16px; }}
    #sortsel {{ padding:12px 14px; }}
    .navi {{ padding:13px 12px; }}
    .bigbtn {{ padding:13px 16px; }}
    .cb-chip {{ padding:7px 13px; }}
    .cb-tab, .cb-btn {{ padding:8px 14px; }}
  }}
</style>
</head>
<body>
{sidebar_html("dash", len(jobs), archived)}
<main>
<div class="wrap">
  <h1>לוח משרות</h1>
  <div class="chipsrow">
    <span class="chip">סה"כ <b>{len(jobs)}</b></span>
    <span class="chip">חדשות <b>{new_count}</b></span>
    <span class="chip mintc">ציון 8+ <b>{top}</b></span>
    <span class="chip pric">הוגשו/ראיון <b>{applied}</b></span>
    {archive_chip}
  </div>
  <div id="batch">
    <button class="bstop" onclick="stopBatch()">עצור</button>
    <span class="bph" id="bphase"></span> <span id="binfo"></span>
    <div id="bprog"><div id="bfill"></div></div>
  </div>
  {focus_html}
  <div class="listbar">
    <input id="q" type="search" placeholder="חיפוש חברה / תפקיד..." aria-label="חיפוש חברה או תפקיד" oninput="applyView()">
    <select id="sortsel" aria-label="מיון הרשימה" onchange="applyView()">
      <option value="score">מיון: ציון</option>
      <option value="date">מיון: הכי חדש</option>
    </select>
    <button class="f on" data-f="הכל" onclick="filt(this)">הכל <span class="cnt">{len(jobs)}</span></button>
    {filters}
  </div>
  <div id="cards">
  {"".join(cards) if cards else empty}
  </div>
</div>
</main>
<div id="cbar">
  <div class="cb-head" onclick="cbToggle()">
    <span class="cb-title">משימות &gt;_</span>
    <span id="cb-chips"><span class="cb-idle">מתחבר לשרת...</span></span>
    <span class="cb-arrow" id="cb-arrow">▲</span>
  </div>
  <div id="cb-panel">
    <div id="cb-tabs"></div>
    <div id="cb-detail"></div>
    <pre id="cb-log" dir="ltr"></pre>
    <div class="cb-actions">
      <button class="cb-btn" onclick="cbCopy()">העתק לוג</button>
      <button class="cb-btn" onclick="cbClear()">נקה משימות</button>
      <button class="cb-btn" id="cb-stop" onclick="cbStop()" style="display:none; border-color:#b91c1c; color:#fca5a5;">עצור משימה</button>
      <label><input type="checkbox" id="cb-auto" checked> גלילה אוטומטית</label>
    </div>
  </div>
</div>
<div class="toast" id="toast" role="status" aria-live="polite"></div>
<script>
// file:// (double-click) talks to cv_server on localhost; served over HTTP
// (phone/LAN/Tailscale) it talks to the same origin that served the page.
const BASE = location.protocol === "file:" ? "http://127.0.0.1:8765" : "";
function toast(msg) {{
  const t = document.getElementById("toast");
  t.textContent = msg;
  t.classList.add("show");
  clearTimeout(t._h);
  t._h = setTimeout(() => t.classList.remove("show"), 4200);
}}
async function serverUp() {{
  try {{
    const h = await fetch(BASE + "/health", {{ signal: AbortSignal.timeout(1500) }});
    return h.ok;
  }} catch (e) {{ return false; }}
}}
function serverDownMsg() {{
  toast("השרת המקומי כבוי - דאבל-קליק על start-cv-server.bat בתיקיית JobHunt ונסה שוב");
}}
// ---- search / sort / filter (one view function) ----
let curFilter = "הכל";
function filt(btn) {{
  document.querySelectorAll(".f").forEach(b => b.classList.remove("on"));
  btn.classList.add("on");
  curFilter = btn.dataset.f;
  applyView();
}}
function applyView() {{
  const q = (document.getElementById("q").value || "").trim().toLowerCase();
  document.querySelectorAll(".card").forEach(c => {{
    const okF = curFilter === "הכל" || c.dataset.status === curFilter;
    const okQ = !q || (c.dataset.search || c.textContent.toLowerCase()).includes(q);
    c.style.display = (okF && okQ) ? "flex" : "none";
  }});
  const by = document.getElementById("sortsel").value;
  const box = document.getElementById("cards");
  const cards = Array.from(box.querySelectorAll(".card"));
  cards.sort((a, b) => by === "date"
    ? (b.dataset.date || "").localeCompare(a.dataset.date || "") || b.dataset.score - a.dataset.score
    : b.dataset.score - a.dataset.score || (b.dataset.date || "").localeCompare(a.dataset.date || ""));
  cards.forEach(c => box.appendChild(c));
}}
// ---- batch pipeline ----
let batchWas = null;
function selIds() {{
  return Array.from(document.querySelectorAll(".selbox:checked")).map(b => b.dataset.id);
}}
function selChanged() {{
  const n = selIds().length;
  const full = document.getElementById("batch-full");
  const cv = document.getElementById("batch-cv");
  if (full) {{ full.textContent = n ? `הגש נבחרות (${{n}})` : "הגש נבחרות"; full.classList.toggle("dim", !n); }}
  if (cv) {{ cv.textContent = n ? `CV לנבחרות (${{n}})` : "CV לנבחרות"; cv.classList.toggle("dim", !n); }}
}}
function subMode() {{
  const el = document.getElementById("submode");
  return el && el.checked ? "auto" : "review";
}}
function applyMode() {{
  const auto = subMode() === "auto";
  document.querySelectorAll(".sub-go").forEach(b => {{
    if (b.disabled) return;
    const idle = b.textContent === "הגש מועמדות" || b.textContent === "הגש אוטומטית";
    if (!idle) return;
    b.textContent = auto ? "הגש אוטומטית" : "הגש מועמדות";
    if (b.dataset.orig) b.dataset.orig = b.textContent;
    b.title = auto
      ? "אוטומטי מלא: Codex ממלא וגם לוחץ Submit בעצמו (יישאל אישור קטן)"
      : "Codex ממלא ועוצר לפני השליחה - אתה בודק בכרום ולוחץ Submit בעצמך";
  }});
}}
async function startBatch(mode) {{
  const ids = selIds();
  if (!ids.length) return toast("סמן קודם משרות עם התיבה בכרטיס - אין הרצה על הכל");
  if (!(await serverUp())) return serverDownMsg();
  const who = ids.length + " המשרות שסימנת";
  const sm = subMode();
  const msg = mode === "full"
    ? (sm === "auto"
        ? "להריץ את הצינור על " + who + "?\\n1) CV למי שאין\\n2) אוטומטי מלא: Codex ימלא ויגיש טופס אחרי טופס לבד\\nעוצר רק על התחברות / CAPTCHA / שאלה שאין לה תשובה."
        : "להריץ את הצינור על " + who + "?\\n1) CV למי שאין\\n2) Codex יכין את כל הטפסים - ממלא כל אחד בטאב משלו ועובר הלאה, בלי לשלוח\\nבסוף מאשרים אחד-אחד מהכרטיסים (אשר ושלח) מתי שנוח - אפשר גם מהטלפון.")
    : "ליצור CV עבור " + who + " (רק למי שאין)? רץ לבד ברקע, אפשר ללכת.";
  if (!confirm(msg)) return;
  const q = "&ids=" + ids.join(",") + (mode === "full" ? "&submit_mode=" + sm : "");
  try {{ await fetch(BASE + "/batch?mode=" + mode + q); }}
  catch (e) {{ return toast("השרת לא מגיב"); }}
  toast(mode === "full" ? "הצינור יצא לדרך" : "יצירת ה-CV התחילה");
  batchPoll();
}}
async function stopBatch() {{
  try {{ await fetch(BASE + "/batch-stop"); toast("ייעצר אחרי המשימה הנוכחית"); }}
  catch (e) {{ toast("השרת לא מגיב"); }}
}}
// ---- on-demand job scan (same routine the scheduled task runs) ----
{SCAN_JS}
function companyOf(id) {{
  const c = document.querySelector("#job-" + id + " .company");
  return c ? c.textContent : "#" + id;
}}
async function batchPoll() {{
  let s;
  try {{ s = await (await fetch(BASE + "/batch-status", {{ signal: AbortSignal.timeout(2500) }})).json(); }}
  catch (e) {{ return; }}
  const el = document.getElementById("batch");
  const running = s.status === "running";
  document.getElementById("batch-full").disabled = running;
  document.getElementById("batch-cv").disabled = running;
  if (!running && s.status !== "done") {{ el.style.display = "none"; return; }}
  if (s.status === "done" && batchWas !== "running") {{ el.style.display = "none"; return; }}
  el.style.display = "block";
  const cvN = (s.cv_done || []).length + (s.cv_failed || []).length;
  const subN = (s.submitted || []).length + (s.ready || []).length +
               (s.blocked || []).length + (s.failed || []).length;
  if (running) {{
    const what = s.phase === "cv" ? "יצירת קורות חיים" : "הגשות עם Codex";
    document.getElementById("bphase").textContent = "הגש הכל · שלב: " + what;
    document.getElementById("binfo").textContent =
      (s.current ? "עכשיו: " + companyOf(s.current) + " (#" + s.current + ")" : "") +
      " · CV: " + cvN + " · הגשות: " + subN + " מתוך " + (s.total || 0);
    const frac = s.total ? (s.phase === "cv" ? cvN / (2 * s.total) : 0.5 + subN / (2 * s.total)) : 0;
    document.getElementById("bfill").style.width = Math.round(frac * 100) + "%";
    setTimeout(batchPoll, 3000);
  }} else {{
    document.getElementById("bphase").textContent = "הצינור הסתיים";
    document.getElementById("binfo").textContent = s.detail || "";
    document.getElementById("bfill").style.width = "100%";
    if (batchWas === "running") toast("הגש הכל הסתיים - " + (s.detail || ""));
  }}
  batchWas = s.status;
}}
// ---- per-job actions ----
async function genCV(id, btn) {{
  if (!(await serverUp())) return serverDownMsg();
  const orig = btn.textContent;
  btn.disabled = true;
  btn.textContent = "מייצר...";
  try {{
    await fetch(BASE + "/generate?id=" + id);
    const t0 = Date.now();
    while (Date.now() - t0 < 30 * 60 * 1000) {{
      await new Promise(r => setTimeout(r, 4000));
      const s = await (await fetch(BASE + "/status?id=" + id)).json();
      if (s.status === "done") {{
        btn.textContent = "מוכן - מרענן...";
        toast("ה-CV מוכן!");
        setTimeout(() => location.reload(), 1500);
        return;
      }}
      if (s.status === "error") {{
        btn.textContent = orig; btn.disabled = false;
        toast("שגיאה: " + String(s.detail || "").slice(0, 140));
        return;
      }}
    }}
    throw new Error("timeout");
  }} catch (e) {{
    btn.textContent = orig; btn.disabled = false;
    toast("נתקע - נסה שוב עוד רגע (או בדוק את חלון השרת)");
  }}
}}
// review package: field summary (+ screenshot) Codex publishes before waiting.
// Shown inside the card so the form can be reviewed and approved from the phone.
function reviewBox(id) {{
  const card = document.getElementById("job-" + id);
  if (!card) return null;
  let box = card.querySelector(".review-box");
  if (!box) {{
    box = document.createElement("div");
    box.className = "review-box";
    box.innerHTML = '<div class="rvh">מה Codex מילא - עבור על הכל, ואז אשר או הגש בכרום בעצמך</div>' +
      '<img class="rvimg" style="display:none" alt="צילום הטופס">' +
      '<pre dir="ltr"></pre>' +
      '<div class="rvb"><button class="btn primary">אשר ושלח</button>' +
      '<button class="btn outline">סגור</button></div>';
    card.querySelector(".body").appendChild(box);
    box.querySelectorAll(".rvb .btn")[1].onclick = () => box.remove();
  }}
  return box;
}}
async function showReview(id, btn, orig) {{
  try {{
    const r = await fetch(BASE + "/review?id=" + id);
    if (!r.ok) return false;
    const s = await r.json();
    const box = reviewBox(id);
    if (!box) return false;
    box.querySelector("pre").textContent = s.summary || "";
    const img = box.querySelector(".rvimg");
    if (s.shot && img.dataset.src !== s.shot) {{
      img.src = BASE + "/" + s.shot; img.dataset.src = s.shot; img.style.display = "block";
      img.onclick = () => window.open(img.src, "_blank");
    }}
    box.querySelectorAll(".rvb .btn")[0].onclick = () => approveSubmit(id, btn, orig);
    return true;
  }} catch (e) {{ return false; }}
}}
async function approveSubmit(id, btn, orig) {{
  if (!confirm("לאשר? Codex ילחץ Submit עכשיו")) return;
  let s;
  try {{ s = await (await fetch(BASE + "/submit-approve?id=" + id)).json(); }}
  catch (e) {{ return toast("השרת לא מגיב"); }}
  if (!s.ok) return toast(s.detail || "האישור נכשל");
  toast("אושר - Codex מגיש עכשיו");
  const box = document.querySelector("#job-" + id + " .review-box");
  if (box) box.querySelectorAll(".rvb .btn")[0].disabled = true;
  if (s.via === "resume") {{
    btn.disabled = true;
    btn.textContent = "Codex מגיש...";
    btn.onclick = null;
    pollSubmit(id, btn, orig);
  }}
}}
// Codex hit a form question with no stored answer: inline box on the card -
// the user answers here (works from the phone too), the server appends it to
// application-answers.md and resumes the same Codex session automatically.
function answerBox(id, question, scroll) {{
  const card = document.getElementById("job-" + id);
  if (!card) return null;
  let box = card.querySelector(".ans-box");
  if (!box) {{
    box = document.createElement("div");
    box.className = "review-box ans-box";
    box.innerHTML = '<div class="rvh">Codex נתקע על שאלה בטופס - ענה כאן והוא ימשיך לבד</div>' +
      '<div class="ans-q" dir="auto"></div>' +
      '<textarea dir="auto" rows="2" placeholder="התשובה שלך (עברית או אנגלית) - נשמרת לתשובות הקבועות לפעמים הבאות"></textarea>' +
      '<div class="rvb"><button class="btn primary">שמור והמשך</button>' +
      '<button class="btn outline">סגור</button></div>';
    card.querySelector(".body").appendChild(box);
    box.querySelectorAll(".rvb .btn")[0].onclick = () => sendAnswer(id, box);
    box.querySelectorAll(".rvb .btn")[1].onclick = () => box.remove();
  }}
  if (question) box.dataset.q = question;
  box.querySelector(".ans-q").textContent =
    box.dataset.q || "(השאלה לא נשמרה - הלוג המלא בקונסול המשימות למטה)";
  if (scroll) box.scrollIntoView({{ block: "nearest" }});
  return box;
}}
async function sendAnswer(id, box) {{
  const answer = box.querySelector("textarea").value.trim();
  if (!answer) return toast("כתוב תשובה קודם");
  if (!(await serverUp())) return serverDownMsg();
  const sv = box.querySelectorAll(".rvb .btn")[0];
  sv.disabled = true; sv.textContent = "שומר...";
  let s;
  try {{
    s = await (await fetch(BASE + "/submit-answer", {{ method: "POST",
      body: JSON.stringify({{ id: id, question: box.dataset.q || "", answer: answer }}) }})).json();
  }} catch (e) {{ sv.disabled = false; sv.textContent = "שמור והמשך"; return toast("השרת לא מגיב"); }}
  if (!s.ok) {{ sv.disabled = false; sv.textContent = "שמור והמשך"; return toast(s.detail || "השמירה נכשלה"); }}
  box.remove();
  const sub = document.querySelector("#job-" + id + " .sub-go");
  if (s.resumed && sub) {{
    toast("נשמר - Codex ממשיך מאיפה שנעצר; Claude מלטש את הניסוח לתשובות הקבועות ברקע");
    const orig = sub.dataset.orig || sub.textContent;
    sub.dataset.orig = orig;
    sub.disabled = true; sub.textContent = "Codex ממשיך..."; sub.onclick = null;
    pollSubmit(id, sub, orig);
  }} else if (s.resumed) {{
    toast("נשמר - Codex ממשיך מאיפה שנעצר; Claude מלטש את הניסוח לתשובות הקבועות ברקע");
  }} else {{
    toast("נשמר - Claude מלטש את הניסוח ברקע; Codex עסוק כרגע, לחץ 'המשך' בכרטיס כשיתפנה");
  }}
}}
// on load: reopen answer boxes for jobs whose Codex run is still waiting on a question
async function loadPending() {{
  try {{
    const r = await fetch(BASE + "/pending", {{ signal: AbortSignal.timeout(2500) }});
    if (!r.ok) return;
    for (const p of ((await r.json()).pending || [])) answerBox(p.id, p.q);
  }} catch (e) {{}}
}}
async function pollSubmit(id, btn, orig) {{
  const restore = () => {{ btn.textContent = orig; btn.disabled = false; btn.onclick = () => submitJob(id, btn); }};
  const contBtn = (msg) => {{
    btn.textContent = "המשך (פתרת בכרום?)";
    btn.disabled = false;
    btn.onclick = () => continueSubmit(id, btn, orig);
    toast(msg);
  }};
  try {{
    const t0 = Date.now();
    while (Date.now() - t0 < 45 * 60 * 1000) {{
      await new Promise(r => setTimeout(r, 5000));
      const s = await (await fetch(BASE + "/submit-status?id=" + id)).json();
      if (s.status === "running") {{ showReview(id, btn, orig); continue; }}
      const d = String(s.detail || "");
      if (s.status === "submitted") {{
        btn.textContent = "הוגש";
        const box = document.querySelector("#job-" + id + " .review-box");
        if (box) box.remove();
        toast("הוגש! Codex זיהה את אישור השליחה - הסטטוס עודכן לבד");
        setTimeout(() => location.reload(), 1600);
        return;
      }}
      if (s.status === "ready") {{
        btn.textContent = "הגשתי - עדכן סטטוס";
        btn.disabled = false;
        btn.onclick = () => setStatus(id, "הוגש", btn);
        await showReview(id, btn, orig);
        toast("הטופס מלא - בדוק את הסיכום בכרטיס ולחץ 'אשר ושלח', או הגש בכרום בעצמך ועדכן סטטוס");
        return;
      }}
      if (s.status === "login") return contBtn("נדרשת התחברות - התחבר בכרום בדף הפתוח, ואז לחץ המשך");
      if (s.status === "captcha") return contBtn("יש CAPTCHA - פתור בכרום, ואז לחץ המשך");
      if (s.status === "needs_input") {{
        answerBox(id, d.replace(/^NEEDS_INPUT:?\\s*/, "").trim(), true);
        return contBtn("Codex נתקע על שאלה - ענה בתיבה שבכרטיס והוא ימשיך לבד");
      }}
      if (s.status === "busy") {{ restore(); toast(d || "הגשה אחרת רצה כרגע - חכה שתסתיים"); return; }}
      restore(); toast("שגיאה: " + d.slice(0, 160));
      return;
    }}
    throw new Error("timeout");
  }} catch (e) {{
    restore();
    toast("נתקע - נסה שוב עוד רגע (או בדוק את הקונסול למטה)");
  }}
}}
async function submitJob(id, btn) {{
  const mode = subMode();
  if (mode === "auto" &&
      !confirm("הגשה אוטומטית מלאה: Codex ימלא וגם ילחץ Submit בלי בדיקה שלך. להמשיך?")) return;
  if (!(await serverUp())) return serverDownMsg();
  const orig = btn.dataset.orig || btn.textContent;
  btn.dataset.orig = orig;
  btn.disabled = true;
  btn.textContent = mode === "auto" ? "Codex מגיש..." : "Codex ממלא...";
  try {{ await fetch(BASE + "/submit?id=" + id + "&mode=" + mode); }}
  catch (e) {{ btn.textContent = orig; btn.disabled = false; return toast("השרת לא מגיב"); }}
  await pollSubmit(id, btn, orig);
}}
async function continueSubmit(id, btn, orig) {{
  if (!(await serverUp())) return serverDownMsg();
  orig = orig || btn.dataset.orig || "הגש מועמדות";
  btn.disabled = true;
  btn.textContent = "Codex ממשיך...";
  try {{ await fetch(BASE + "/submit-continue?id=" + id); }}
  catch (e) {{ btn.textContent = orig; btn.disabled = false; return toast("השרת לא מגיב"); }}
  await pollSubmit(id, btn, orig);
}}
async function setStatus(id, status, el) {{
  if (!(await serverUp())) return serverDownMsg();
  if (el) el.disabled = true;
  try {{
    const r = await fetch(BASE + "/set-status?id=" + id + "&status=" + encodeURIComponent(status),
                          {{ signal: AbortSignal.timeout(30000) }});
    const s = await r.json();
    if (!s.ok) throw new Error(s.detail || "server error");
    toast("סטטוס עודכן: " + status);
    setTimeout(() => location.reload(), 700);
  }} catch (e) {{
    if (el) el.disabled = false;
    toast("עדכון נכשל: " + String(e.message || "").slice(0, 120));
  }}
}}
async function delJob(id, btn) {{
  if (!confirm("למחוק את משרה #" + id + " מהטבלה?")) return;
  if (!(await serverUp())) return serverDownMsg();
  btn.disabled = true;
  try {{
    const r = await fetch(BASE + "/delete?id=" + id, {{ signal: AbortSignal.timeout(30000) }});
    const s = await r.json();
    if (!s.ok) throw new Error(s.detail || "server error");
    toast("נמחקה משרה #" + id);
    setTimeout(() => location.reload(), 700);
  }} catch (e) {{
    btn.disabled = false;
    toast("מחיקה נכשלה: " + String(e.message || "").slice(0, 120));
  }}
}}
// ---- task console ----
let cbOpen = false, cbSel = null, cbFrom = 0, cbTasks = [], cbPrev = {{}};
let cbNow = Date.now() / 1000, cbDash = null;
const CB_STATUS = {{
  running: ["רץ", "running"], done: ["הסתיים", "ok"], error: ["שגיאה", "err"],
  ready: ["ממתין ל-Submit שלך", "warn"], submitted: ["הוגש", "ok"],
  login: ["צריך התחברות", "warn"], captcha: ["צריך CAPTCHA", "warn"],
  needs_input: ["חסרה תשובה", "warn"], busy: ["עסוק", "warn"],
  stopped: ["נעצר", "warn"],
}};
const CB_KIND = {{ claude: "Claude", codex: "Codex", batch: "צינור" }};
function cbFmt(t, now) {{
  const s = Math.max(0, Math.round((t.ended || now) - t.started));
  return String(Math.floor(s / 60)).padStart(2, "0") + ":" + String(s % 60).padStart(2, "0");
}}
function cbToggle() {{
  cbOpen = !cbOpen;
  document.getElementById("cb-panel").style.display = cbOpen ? "block" : "none";
  document.getElementById("cb-arrow").textContent = cbOpen ? "▼" : "▲";
  if (cbOpen && !cbSel && cbTasks.length) cbSelect(cbTasks[0].key);
  else cbRender();
}}
function cbSelect(key) {{
  if (cbSel !== key) {{
    cbSel = key; cbFrom = 0;
    document.getElementById("cb-log").textContent = "";
    document.getElementById("cb-detail").style.display = "none";
  }}
  if (!cbOpen) {{
    cbOpen = true;
    document.getElementById("cb-panel").style.display = "block";
    document.getElementById("cb-arrow").textContent = "▼";
  }}
  cbRender();
  cbFetchLog();
}}
function cbLabel(t) {{
  const tag = t.kind === "batch" ? "" : " #" + t.job_id;
  return (CB_KIND[t.kind] || t.kind) + tag;
}}
function cbRender() {{
  const chips = document.getElementById("cb-chips");
  if (!cbTasks.length) {{
    chips.innerHTML = '<span class="cb-idle">אין משימות עדיין - ריצות של Claude ו-Codex יופיעו כאן</span>';
  }} else {{
    chips.innerHTML = cbTasks.slice(0, 4).map(t => {{
      const st = CB_STATUS[t.status] || [t.status, ""];
      const spin = t.status === "running" ? '<span class="cb-spin">◐</span> ' : "";
      return `<span class="cb-chip ${{st[1]}}" onclick="event.stopPropagation(); cbSelect('${{t.key}}')">${{spin}}${{cbLabel(t)}} · ${{st[0]}} · ${{cbFmt(t, cbNow)}}</span>`;
    }}).join("");
  }}
  if (!cbOpen) return;
  document.getElementById("cb-tabs").innerHTML = cbTasks.map(t => {{
    const st = CB_STATUS[t.status] || [t.status, ""];
    return `<button class="cb-tab${{t.key === cbSel ? " sel" : ""}}" onclick="cbSelect('${{t.key}}')">${{CB_KIND[t.kind] || t.kind}} · ${{t.label}} · ${{st[0]}}</button>`;
  }}).join("");
  const sel = cbTasks.find(t => t.key === cbSel);
  document.getElementById("cb-stop").style.display =
    (sel && sel.status === "running") ? "inline-block" : "none";
}}
async function cbClear() {{
  try {{
    const r = await fetch(BASE + "/tasks-clear", {{ signal: AbortSignal.timeout(5000) }});
    const j = await r.json();
    cbSel = null; cbFrom = 0;
    document.getElementById("cb-log").textContent = "";
    document.getElementById("cb-detail").style.display = "none";
    toast(j.removed ? "נוקו " + j.removed + " משימות" : "אין משימות שהסתיימו לנקות");
  }} catch (e) {{
    toast("ניקוי נכשל - השרת לא מגיב");
  }}
}}
async function cbStop() {{
  const sel = cbTasks.find(t => t.key === cbSel);
  if (!sel) return;
  if (!confirm("לעצור את: " + sel.label + "?")) return;
  try {{
    const r = await fetch(BASE + "/task-stop?key=" + encodeURIComponent(cbSel),
                          {{ signal: AbortSignal.timeout(20000) }});
    const j = await r.json();
    toast(j.ok ? "המשימה נעצרה" : "עצירה נכשלה: " + j.detail);
  }} catch (e) {{
    toast("עצירה נכשלה - השרת לא מגיב");
  }}
}}
async function cbPoll() {{
  let delay = 5000;
  try {{
    const r = await fetch(BASE + "/tasks", {{ signal: AbortSignal.timeout(2500) }});
    const j = await r.json();
    cbTasks = j.tasks || [];
    cbNow = j.now || Date.now() / 1000;
    if (j.dash_mtime) {{
      if (cbDash === null) {{
        cbDash = j.dash_mtime;
      }} else if (j.dash_mtime !== cbDash) {{
        cbDash = j.dash_mtime;
        toast("הדשבורד התעדכן - מרענן...");
        setTimeout(() => location.reload(), 900);
        return;
      }}
    }}
    for (const t of cbTasks) {{
      if (cbPrev[t.key] === "running" && t.status === "error") {{
        cbSelect(t.key);
        toast("משימה נכשלה - פרטים בקונסול למטה");
      }}
      cbPrev[t.key] = t.status;
    }}
    if (cbTasks.some(t => t.status === "running")) delay = 2500;
    cbRender();
    if (cbOpen && cbSel) await cbFetchLog();
  }} catch (e) {{
    document.getElementById("cb-chips").innerHTML =
      '<span class="cb-idle">השרת כבוי - start-cv-server.bat</span>';
    delay = 8000;
  }}
  setTimeout(cbPoll, delay);
}}
async function cbFetchLog() {{
  if (!cbSel) return;
  try {{
    const r = await fetch(BASE + "/task-log?key=" + encodeURIComponent(cbSel) + "&from=" + cbFrom,
                          {{ signal: AbortSignal.timeout(2500) }});
    if (!r.ok) return;
    const j = await r.json();
    const log = document.getElementById("cb-log");
    if (j.lines && j.lines.length) {{
      log.textContent += j.lines.join("\\n") + "\\n";
      cbFrom = j.next;
      if (document.getElementById("cb-auto").checked) log.scrollTop = log.scrollHeight;
    }}
    const det = document.getElementById("cb-detail");
    const isErr = j.status === "error";
    const isWarn = ["ready", "login", "captcha", "needs_input", "busy"].includes(j.status);
    if (j.detail && (isErr || isWarn)) {{
      det.style.display = "block";
      det.className = isWarn ? "warn" : "";
      det.textContent = j.detail;
      const selTask = cbTasks.find(t => t.key === cbSel);
      if (selTask && selTask.kind === "codex" &&
          ["login", "captcha", "needs_input"].includes(j.status)) {{
        const b = document.createElement("button");
        b.className = "cb-btn";
        b.style.marginRight = "10px";
        b.textContent = "המשך את Codex";
        b.onclick = () => {{ b.disabled = true; continueSubmit(selTask.job_id, b, "המשך את Codex"); }};
        det.appendChild(b);
      }}
    }} else {{
      det.style.display = "none";
    }}
  }} catch (e) {{ /* retry on next poll */ }}
}}
function cbCopy() {{
  const det = document.getElementById("cb-detail");
  const txt = (det.style.display !== "none" ? det.textContent + "\\n\\n" : "") +
              document.getElementById("cb-log").textContent;
  const done = () => toast("הלוג הועתק - אפשר להדביק לי בצ'אט לאבחון");
  if (navigator.clipboard && window.isSecureContext) {{
    navigator.clipboard.writeText(txt).then(done).catch(() => cbCopyFallback(txt, done));
  }} else {{
    cbCopyFallback(txt, done);
  }}
}}
function cbCopyFallback(txt, done) {{
  const ta = document.createElement("textarea");
  ta.value = txt;
  ta.style.position = "fixed"; ta.style.opacity = "0";
  document.body.appendChild(ta);
  ta.focus(); ta.select();
  try {{
    document.execCommand("copy") ? done() : toast("העתקה נכשלה - סמן ידנית מהלוג");
  }} catch (e) {{
    toast("העתקה נכשלה - סמן ידנית מהלוג");
  }}
  document.body.removeChild(ta);
}}
cbPoll();
selChanged();
batchPoll();
scanPoll();
loadPending();
(() => {{
  const el = document.getElementById("submode");
  if (!el) return;
  el.checked = localStorage.getItem("submode") === "auto";
  el.onchange = () => {{
    localStorage.setItem("submode", el.checked ? "auto" : "review");
    applyMode();
  }};
  applyMode();
}})();
// prepared forms survive a page reload: any job with a published review
// package gets its box (סיכום + אשר ושלח) restored on load
(async () => {{
  for (const b of document.querySelectorAll(".sub-go")) {{
    const card = b.closest(".card");
    if (!card) continue;
    const id = parseInt(card.id.replace("job-", ""), 10);
    if (id) await showReview(id, b, b.textContent);
  }}
}})();
</script>
</body>
</html>"""


# ---------------------------------------------------------------- sub-pages
SUB_CSS = """
  :root {
    --bg:#f8fafc; --card:#ffffff; --line:#e2e8f0; --line2:#cbd5e1; --ink:#091426; --mut:#45474c;
    --navy:#091426; --navy-h:#1e293b; --navy-ink:#ffffff;
    --pri:#712ae2; --pri-h:#5e1fca; --pri-ink:#ffffff;
    --mint:#6ffbbe; --mint-dim:#4edea3; --mint-ink:#002113;
    --gray-badge:#eae7e9; --soft:#f0edef; --softer:#f5f3f4;
    --ok:#1a7f37; --okbg:#e6f6ea; --err:#ba1a1a; --errbg:#ffdad6;
    --shadow:0 4px 12px rgba(30,41,59,.05);
  }
  * { margin:0; padding:0; box-sizing:border-box; }
  html { color-scheme: light; }
  body { font-family:"Work Sans","Segoe UI",Arial,sans-serif; background:var(--bg); color:var(--ink); }
  .ic { width:1em; height:1em; vertical-align:-0.13em; flex-shrink:0; }
  :focus-visible { outline:2px solid var(--pri); outline-offset:2px; }
  aside { position:fixed; top:0; right:0; bottom:0; width:250px; z-index:50;
          background:var(--card); border-left:1px solid var(--line);
          display:flex; flex-direction:column; padding:18px 14px;
          box-shadow:0 0 12px rgba(30,41,59,.04); }
  .brand { font-size:17px; font-weight:800; letter-spacing:-.02em; padding:2px 10px 14px; }
  .userbox { display:flex; flex-direction:column; gap:2px; background:var(--softer);
             border:1px solid var(--line); border-radius:12px; padding:10px 12px; margin-bottom:16px; }
  .uname { font-weight:700; font-size:14px; }
  .urole { color:var(--mut); font-size:12px; }
  nav.side { display:flex; flex-direction:column; gap:4px; }
  .navi { display:flex; align-items:center; gap:10px; padding:10px 12px; border-radius:12px;
          text-decoration:none; color:var(--mut); font-size:13.5px; font-weight:600;
          transition:background .12s; }
  .navi:hover { background:var(--softer); }
  .navi.act { background:var(--pri); color:var(--pri-ink); }
  .navi .cnt { margin-right:auto; background:var(--soft); border-radius:9999px; padding:1px 8px;
               font-size:12px; color:var(--mut); }
  .navi.act .cnt { background:rgba(255,255,255,.2); color:var(--pri-ink); }
  .sidegrow { flex:1; }
  .sideact { display:flex; flex-direction:column; gap:8px; }
  .bigbtn { border:none; border-radius:10px; padding:11px 16px; font-size:13.5px; font-weight:700;
            cursor:pointer; font-family:inherit; background:var(--pri); color:var(--pri-ink);
            box-shadow:var(--shadow); transition:transform .08s; width:100%; }
  .bigbtn:hover { background:var(--pri-h); }
  .bigbtn:active { transform:scale(.98); }
  .bigbtn.sec { background:var(--card); color:var(--ink); border:1px solid var(--line2); box-shadow:none; }
  .bigbtn.sec:hover { background:var(--softer); }
  .bigbtn:disabled { opacity:.55; cursor:default; }
  .bigbtn.dim { opacity:.55; }
  .swrow { display:flex; align-items:center; gap:8px; cursor:pointer; padding:0 2px 2px; user-select:none; }
  .swrow input { display:none; }
  .sw { width:34px; height:19px; border-radius:10px; background:var(--line2); position:relative;
        transition:background .15s; flex:none; }
  .sw::after { content:""; position:absolute; top:2px; right:2px; width:15px; height:15px; border-radius:50%;
               background:#fff; transition:transform .15s; box-shadow:0 1px 2px rgba(0,0,0,.25); }
  .swrow input:checked ~ .sw { background:var(--pri); }
  .swrow input:checked ~ .sw::after { transform:translateX(-15px); }
  .swtxt { font-size:12.5px; font-weight:600; color:var(--mut); display:flex; align-items:center; gap:6px; }
  .swrow input:checked ~ .swtxt { color:var(--pri); }
  .sidefoot { color:var(--mut); font-size:12px; line-height:1.7; padding:12px 10px 0; }
  main { margin-right:250px; padding:22px 26px 60px; }
  .wrap { max-width:900px; margin:0 auto; }
  h1 { font-size:24px; font-weight:700; letter-spacing:-.02em; margin-bottom:4px; }
  .sub { color:var(--mut); font-size:12.5px; margin-bottom:16px; }
  .panel { background:var(--card); border:1px solid var(--line); border-radius:12px;
           padding:16px 18px; margin-bottom:14px; box-shadow:var(--shadow); }
  .panel h2 { font-size:16px; font-weight:700; margin-bottom:10px; }
  .panel h3 { font-size:14px; font-weight:600; margin:12px 0 6px; }
  .panel p, .panel li { font-size:13.5px; line-height:1.65; color:var(--ink); }
  .panel ul { padding-right:20px; }
  .mut { color:var(--mut); }
  .note { background:#eef6ff; border:1px solid #cfe3f7; border-radius:10px; padding:10px 14px;
          font-size:13px; margin-bottom:16px; }
  .toast { position:fixed; bottom:24px; right:50%; transform:translateX(50%); background:var(--navy);
           color:var(--navy-ink); padding:10px 22px; border-radius:10px; font-size:14px; opacity:0;
           transition:opacity .25s; pointer-events:none; z-index:70; }
  .toast.show { opacity:1; }
  @media (prefers-reduced-motion:reduce) {
    *, *::before, *::after { transition:none !important; animation:none !important; }
  }
  @media (max-width:900px) {
    aside { position:static; width:auto; border-left:none; border-bottom:1px solid var(--line);
            flex-direction:row; flex-wrap:wrap; align-items:center; gap:8px; padding:12px 16px; }
    .brand { padding:0 0 0 6px; }
    .userbox { flex-direction:row; gap:8px; align-items:baseline; margin:0; padding:6px 12px; }
    nav.side { flex-direction:row; flex-wrap:wrap; }
    .sidegrow { display:none; }
    .sideact { flex-direction:row; }
    .bigbtn { width:auto; }
    .sidefoot { display:none; }
    main { margin-right:0; padding:16px 14px 60px; }
  }
  @media (pointer:coarse) {
    .btn { padding:12px 15px; }
    .btn.iconb { padding:12px 12px; }
    .navi { padding:13px 12px; }
    .bigbtn { padding:13px 16px; }
    #q { padding:12px 16px; }
  }
"""


def sidebar_html(active, jobs_n, archived_n):
    """Shared fixed sidebar - identical on every page so it never 'disappears'."""
    try:
        cv = json.loads((ROOT / "master-cv.json").read_text(encoding="utf-8-sig"))
        name, role = cv.get("name", "לידור"), cv.get("title", "")
    except Exception:  # noqa: BLE001
        name, role = "לידור", ""
    def cls(k):
        return "navi act" if k == active else "navi"
    return f"""<aside>
  <div class="brand">JobHunt</div>
  <div class="userbox">
    <span class="uname">{html.escape(name)}</span>
    <span class="urole">{html.escape(role)}</span>
  </div>
  <nav class="side">
    <a class="{cls('dash')}" href="dashboard.html">{IC_GRID} לוח משרות <span class="cnt">{jobs_n}</span></a>
    <a class="{cls('archive')}" href="archive.html">{IC_BOX} ארכיון <span class="cnt">{archived_n}</span></a>
    <a class="{cls('profile')}" href="profile.html">{IC_USER} הפרופיל שלי</a>
  </nav>
  <div class="sidegrow"></div>
  <div class="sideact">
    <label class="swrow" title="כבוי: Codex ממלא ועוצר - אתה בודק בכרום ולוחץ Submit. דלוק: Codex גם מגיש בעצמו. חל על כפתור ההגשה בכרטיס וגם על הגש נבחרות">
      <input type="checkbox" id="submode">
      <span class="sw"></span>
      <span class="swtxt">{IC_BOLT} Codex מגיש לבד</span>
    </label>
    <button class="bigbtn dim" id="batch-full" onclick="startBatch('full')"
            title="רק למשרות שסימנת עם התיבה בכרטיס: יוצר CV למי שאין, ואז Codex ממלא טופס אחרי טופס - מצב השליחה לפי המתג שלמעלה">הגש נבחרות</button>
    <button class="bigbtn sec dim" id="batch-cv" onclick="startBatch('cv')"
            title="יצירת CV רק למשרות שסימנת (למי שאין) - רץ לבד, אפשר ללכת">CV לנבחרות</button>
    <div class="scangrp">
      <div class="scanh">{IC_SEARCH} סריקת משרות - הרץ עכשיו</div>
      <div class="scanrow">
        <button class="bigbtn scan" id="scan-cheap" onclick="runScan('cheap')"
                title="הסריקה היומית: NVIDIA / Intel / Amazon / לינקדאין (חיפה והצפון) / ג'וב מאסטר, רק משרות מ-72 השעות האחרונות. כמה דקות.">מהירה</button>
        <button class="bigbtn sec scan" id="scan-full" onclick="runScan('full')"
                title="סריקה מלאה: כל המקורות שנגישים ב-HTTP, חלון של 7 ימים. ארוכה יותר. מקורות שדורשים דפדפן (מיקרוסופט/גוגל/אפל/צ'ק פוינט/דרושים/אולג'ובס/וואטסאפ) לא רצים מכאן - הם נשארים לסריקה המתוזמנת.">מלאה</button>
      </div>
      <div class="scanst" id="scanst"></div>
    </div>
  </div>
  <div class="sidefoot">עודכן {date.today().isoformat()}<br>ארכוב אוטומטי אחרי {STALE_NEW_DAYS} יום</div>
</aside>"""


# Sub-pages have no job checkboxes - batch runs are selection-only, so the
# sidebar buttons just send the user to the dashboard to pick jobs there.
SIDE_JS = """
const BASE = location.protocol === "file:" ? "http://127.0.0.1:8765" : "";
function toast(m) {
  const t = document.getElementById("toast");
  t.textContent = m; t.classList.add("show");
  clearTimeout(t._h); t._h = setTimeout(() => t.classList.remove("show"), 4000);
}
function startBatch(mode) {
  toast("הבחירה נעשית בלוח - סמן שם משרות ואז הפעל");
  setTimeout(() => location.href = "dashboard.html", 900);
}
""" + SCAN_JS + """
scanPoll();
(() => {
  const el = document.getElementById("submode");
  if (!el) return;
  el.checked = localStorage.getItem("submode") === "auto";
  el.onchange = () => localStorage.setItem("submode", el.checked ? "auto" : "review");
})();
"""


def load_archive():
    wb = load_workbook(TRACKER)
    if "Archive" not in wb.sheetnames:
        return []
    rows = []
    for r in wb["Archive"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        rows.append({
            "id": r[0], "date": str(r[1] or "")[:10], "company": r[2] or "",
            "title": r[3] or "", "location": r[4] or "", "score": r[5] or 0,
            "reason": r[6] or "", "link": r[7] or "", "status": r[9] or "",
            "arc_date": str(r[13])[:10] if len(r) > 13 and r[13] else "",
            "arc_reason": r[14] if len(r) > 14 and r[14] else "",
        })
    rows.sort(key=lambda x: (x["arc_date"], x["id"]), reverse=True)
    return rows


def build_archive(rows, jobs_n=0):
    cards = []
    for j in rows:
        link_a = (f'<a class="btn outline iconb" href="{html.escape(str(j["link"]), quote=True)}" '
                  f'target="_blank" title="פתיחת המשרה במקור" aria-label="פתיחת המשרה במקור">{IC_LINK}</a>') if j["link"] else ""
        blob = " ".join(str(j[k]) for k in ("company", "title", "location", "arc_reason", "status"))
        blob = f'{blob} #{j["id"]}'.lower()
        cards.append(f"""
    <div class="acard" id="ar-{j['id']}" data-search="{html.escape(blob, quote=True)}">
      <span class="fscore s{min(int(j['score'] or 0), 10)}">{j['score']}</span>
      <div class="atxt">
        <div class="arow"><span class="ac">{html.escape(str(j['company']))}</span>
          <span class="meta">{html.escape(str(j['location']))}</span>
          <span class="meta mut">פורסם {html.escape(j['date'])}</span>
          <span class="meta mut">#{j['id']}</span></div>
        <div class="at">{html.escape(str(j['title']))}</div>
        <div class="aw">{html.escape(str(j['arc_reason']))} · נכנס לארכיון {html.escape(j['arc_date'])} · היה בסטטוס "{html.escape(str(j['status']))}"</div>
      </div>
      <div class="aact">{link_a}
        <button class="btn navy" onclick="restoreJob({j['id']}, this)">שחזר ללוח</button>
        <button class="btn del iconb" onclick="delArchived({j['id']}, this)" title="מחיקה סופית מהארכיון (לא יחזור בסריקות)" aria-label="מחיקה סופית מהארכיון">{IC_TRASH}</button></div>
    </div>""")
    body = "".join(cards) if cards else '<div class="panel mut">הארכיון ריק.</div>'
    clear_btn = (f'<button class="btn delall" onclick="clearArchive()">מחק הכל ({len(rows)})</button>'
                 if rows else "")
    return f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>JobHunt - ארכיון</title>
<style>
{SUB_CSS}
  .acard {{ display:flex; gap:12px; align-items:flex-start; background:var(--card);
            border:1px solid var(--line); border-radius:12px; padding:13px 15px;
            margin-bottom:10px; box-shadow:var(--shadow); }}
  .fscore {{ font-weight:700; font-size:14px; border-radius:8px; padding:3px 9px;
             background:var(--gray-badge); }}
  .fscore.s10,.fscore.s9 {{ background:var(--mint); color:var(--mint-ink); }}
  .fscore.s8 {{ background:var(--mint-dim); color:var(--mint-ink); }}
  .atxt {{ flex:1; min-width:0; }}
  .arow {{ display:flex; gap:10px; flex-wrap:wrap; align-items:baseline; }}
  .ac {{ font-weight:700; font-size:14px; }}
  .meta {{ color:var(--mut); font-size:12px; }}
  .at {{ font-size:14.5px; font-weight:600; margin:3px 0; }}
  .aw {{ color:var(--mut); font-size:12px; }}
  .aact {{ display:flex; gap:8px; align-items:center; }}
  .btn {{ border:none; border-radius:9px; padding:7px 14px; font-size:12.5px; font-weight:600;
          cursor:pointer; text-decoration:none; display:inline-flex; align-items:center; gap:5px;
          font-family:inherit; }}
  .btn.navy {{ background:var(--navy); color:var(--navy-ink); }}
  .btn.navy:hover {{ background:var(--navy-h); }}
  .btn.outline {{ background:var(--card); color:var(--ink); border:1px solid var(--line); }}
  .btn.iconb {{ padding:7px 9px; font-size:14px; }}
  .btn.del {{ background:none; color:var(--err); }}
  .btn.del:hover {{ background:var(--errbg); }}
  .btn.delall {{ background:none; color:var(--err); border:1px solid var(--err); }}
  .btn.delall:hover {{ background:var(--errbg); }}
  .btn:disabled {{ opacity:.6; cursor:default; }}
  .abar {{ display:flex; gap:10px; align-items:center; flex-wrap:wrap; margin-bottom:14px; }}
  #q {{ border:1px solid var(--line); background:var(--card); color:var(--ink);
        border-radius:9999px; padding:9px 14px; font-size:13px; font-family:inherit; width:240px; }}
  @media (pointer:coarse) {{
    .acard .btn {{ padding:12px 15px; }}
    .acard .btn.iconb {{ padding:12px 12px; }}
  }}
</style>
</head>
<body>
{sidebar_html("archive", jobs_n, len(rows))}
<main>
<div class="wrap">
  <h1>ארכיון משרות</h1>
  <div class="sub">{len(rows)} משרות. שחזור מחזיר ללוח בסטטוס "חדש" (ולא ייארכב שוב); מחיקה היא סופית - המשרה לא תחזור בסריקות.</div>
  <div class="abar">
    <input id="q" type="search" placeholder="חיפוש בארכיון..." aria-label="חיפוש בארכיון" oninput="flt()">
    {clear_btn}
  </div>
  <div id="list">{body}</div>
</div>
</main>
<div class="toast" id="toast" role="status" aria-live="polite"></div>
<script>
{SIDE_JS}
function flt() {{
  const q = (document.getElementById("q").value || "").trim().toLowerCase();
  document.querySelectorAll(".acard").forEach(c => {{
    c.style.display = (!q || (c.dataset.search || c.textContent.toLowerCase()).includes(q)) ? "flex" : "none";
  }});
}}
async function restoreJob(id, btn) {{
  btn.disabled = true; btn.textContent = "משחזר...";
  try {{
    const r = await fetch(BASE + "/restore?id=" + id, {{ signal: AbortSignal.timeout(30000) }});
    const s = await r.json();
    if (!s.ok) throw new Error(s.detail || "server error");
    toast("שוחזר ללוח");
    setTimeout(() => location.reload(), 900);
  }} catch (e) {{
    btn.disabled = false; btn.textContent = "שחזר ללוח";
    toast("שחזור נכשל: " + String(e.message || "").slice(0, 120) +
          " (השרת רץ? start-cv-server.bat)");
  }}
}}
async function delArchived(id, btn) {{
  if (!confirm("למחוק סופית את משרה #" + id + " מהארכיון? לא תחזור בסריקות.")) return;
  btn.disabled = true;
  try {{
    const r = await fetch(BASE + "/archive-delete?id=" + id, {{ signal: AbortSignal.timeout(30000) }});
    const s = await r.json();
    if (!s.ok) throw new Error(s.detail || "server error");
    toast("נמחקה מהארכיון");
    setTimeout(() => location.reload(), 900);
  }} catch (e) {{
    btn.disabled = false;
    toast("מחיקה נכשלה: " + String(e.message || "").slice(0, 120));
  }}
}}
async function clearArchive() {{
  const n = document.querySelectorAll(".acard").length;
  if (!confirm("למחוק סופית את כל " + n + " המשרות מהארכיון? הן לא יחזרו בסריקות.")) return;
  try {{
    const r = await fetch(BASE + "/archive-clear", {{ signal: AbortSignal.timeout(60000) }});
    const s = await r.json();
    if (!s.ok) throw new Error(s.detail || "server error");
    toast("הארכיון נוקה (" + (s.deleted || 0) + " נמחקו)");
    setTimeout(() => location.reload(), 900);
  }} catch (e) {{
    toast("מחיקה נכשלה: " + String(e.message || "").slice(0, 120) +
          " (השרת רץ? start-cv-server.bat)");
  }}
}}
</script>
</body>
</html>"""


def render_md(text):
    """Tiny markdown -> HTML for profile.md (headings, bullets, bold)."""
    import re as _re
    out, in_ul = [], False
    for line in text.splitlines():
        s = html.escape(line.rstrip())
        s = _re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", s)
        if s.startswith("- "):
            if not in_ul:
                out.append("<ul>")
                in_ul = True
            out.append(f"<li>{s[2:]}</li>")
            continue
        if in_ul:
            out.append("</ul>")
            in_ul = False
        if s.startswith("### "):
            out.append(f"<h3>{s[4:]}</h3>")
        elif s.startswith("## "):
            out.append(f"<h3>{s[3:]}</h3>")
        elif s.startswith("# "):
            out.append(f"<h2>{s[2:]}</h2>")
        elif s:
            out.append(f"<p>{s}</p>")
    if in_ul:
        out.append("</ul>")
    return "\n".join(out)


def build_profile(jobs_n=0, archived_n=0):
    cv = json.loads((ROOT / "master-cv.json").read_text(encoding="utf-8-sig"))
    prof_md = (ROOT / "profile.md").read_text(encoding="utf-8-sig")

    def esc(x):
        return html.escape(str(x))

    exp_html = ""
    for e in cv.get("experience", []):
        bullets = "".join(f"<li>{esc(b)}</li>" for b in e.get("bullets", []))
        note = f'<p class="mut">{esc(e["note"])}</p>' if e.get("note") else ""
        exp_html += (f'<h3>{esc(e.get("role", ""))} | {esc(e.get("company", ""))} '
                     f'<span class="mut">({esc(e.get("period", ""))})</span></h3>{note}<ul>{bullets}</ul>')
    edu_html = ""
    for e in cv.get("education", []):
        hi = "".join(f"<li>{esc(h)}</li>" for h in e.get("highlights", []))
        edu_html += (f'<h3>{esc(e.get("degree", ""))} | {esc(e.get("institution", ""))} '
                     f'<span class="mut">({esc(e.get("period", ""))})</span></h3><ul>{hi}</ul>')
    proj_html = ""
    for p in cv.get("projects", []) + cv.get("projects_pool", []):
        bl = "".join(f"<li>{esc(b)}</li>" for b in p.get("bullets", []))
        proj_html += f'<h3>{esc(p.get("name", ""))} <span class="mut">({esc(p.get("tech", ""))})</span></h3><ul>{bl}</ul>'
    skills = cv.get("skills", {})
    skills_html = "".join(
        f'<p><b>{esc(k)}:</b> {esc(", ".join(v))}</p>' for k, v in skills.items())
    awards = "".join(f"<li>{esc(a)}</li>" for a in cv.get("awards", []))
    contact = cv.get("contact", {})

    return f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>JobHunt - הפרופיל שלי</title>
<style>
{SUB_CSS}
</style>
</head>
<body>
{sidebar_html("profile", jobs_n, archived_n)}
<main>
<div class="wrap">
  <h1>מה ה-AI יודע עליי</h1>
  <div class="sub">משני קבצים: master-cv.json (תוכן קורות החיים - מקור האמת לכל CV שנוצר) + profile.md (קריטריוני החיפוש והציונים)</div>
  <div class="note">מצאת משהו לא מדויק או חסר? תגיד ל-Claude בצ'אט (למשל: "תעדכן את ה-CV הראשי: ...") - זה מתעדכן בשני הקבצים והדף הזה נבנה מחדש.</div>

  <div class="panel">
    <h2>{esc(cv.get('name', ''))} - {esc(cv.get('title', ''))}</h2>
    <p class="mut">{esc(contact.get('location', ''))} · {esc(contact.get('phone', ''))} · {esc(contact.get('email', ''))}</p>
    <p class="mut">LinkedIn: {esc(contact.get('linkedin_label', ''))} · GitHub: {esc(contact.get('github_label', ''))}</p>
    <h3>תקציר</h3>
    <p>{esc(cv.get('summary_base', ''))}</p>
  </div>

  <div class="panel"><h2>ניסיון תעסוקתי</h2>{exp_html}</div>
  <div class="panel"><h2>השכלה והצטיינות</h2>{edu_html}<ul>{awards}</ul></div>
  <div class="panel"><h2>פרויקטים (כולל מאגר להחלפה לפי משרה)</h2>{proj_html}</div>
  <div class="panel"><h2>כישורים</h2>{skills_html}
    <p><b>שירות צבאי:</b> {esc(cv.get('military', {}).get('text', ''))}</p>
    <p><b>שפות:</b> {esc(', '.join(cv.get('languages_spoken', [])))}</p></div>

  <div class="panel"><h2>קריטריוני החיפוש (profile.md)</h2>{render_md(prof_md)}</div>
</div>
</main>
<div class="toast" id="toast" role="status" aria-live="polite"></div>
<script>
{SIDE_JS}
</script>
</body>
</html>"""


def main():
    try:
        n, moved = archive_stale.run()
        if n:
            print(f"auto-archived {n}: " + "; ".join(moved))
    except Exception as e:  # noqa: BLE001 - Excel holding the file must not kill the rebuild
        print(f"archive pass skipped: {e}")
    jobs, archived = load_jobs()
    OUT.write_text(build(jobs, archived), encoding="utf-8")
    (ROOT / "archive.html").write_text(build_archive(load_archive(), len(jobs)), encoding="utf-8")
    try:
        (ROOT / "profile.html").write_text(build_profile(len(jobs), archived), encoding="utf-8")
    except Exception as e:  # noqa: BLE001 - a malformed master-cv must not kill the rebuild
        print(f"profile page skipped: {e}")
    print(f"dashboard: {OUT} ({len(jobs)} jobs, {archived} archived) + archive.html + profile.html")


if __name__ == "__main__":
    main()
