# -*- coding: utf-8 -*-
"""Build report.html - a read-only, phone-friendly copy of the tracker for publishing
as a private Claude artifact (so the list is reachable from any device).

Usage:
    python make_web_report.py            # writes ROOT/report.html

Unlike dashboard.html this page has no Generate-CV / delete / status buttons: those
drive local scripts and cannot work off this machine. Re-run after every scan, then
re-publish the artifact to the SAME url.
"""
import io
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TRACKER = ROOT / "jobs.xlsx"
OUT = ROOT / "report.html"


def read_jobs():
    ws = load_workbook(TRACKER)["Jobs"]
    jobs = []
    for r in ws.iter_rows(min_row=2):
        v = [c.value for c in r]
        if v[0] is None:
            continue
        link = r[7].hyperlink.target if r[7].hyperlink else v[7]
        jobs.append({"n": v[0], "date": v[1] or "", "co": v[2] or "", "title": v[3] or "",
                     "loc": v[4] or "", "score": v[5] or 0, "reason": v[6] or "",
                     "link": link or "", "status": v[9] or "חדש", "cv": v[10] or ""})
    jobs.sort(key=lambda j: (-(j["score"] or 0), str(j["date"])), reverse=False)
    return jobs


HTML = u"""<title>משרות - לידור בן חמו</title>
<style>
  :root {
    --ink:#16385E; --ink-2:#42556E; --muted:#6B7A8D; --accent:#2563A8;
    --good:#1A7F37; --warn:#B26A00; --ground:#F2F5F9; --card:#FFFFFF;
    --line:#DFE6EE; --chip:#EEF3F9;
  }
  @media (prefers-color-scheme: dark) {
    :root {
      --ink:#D8E4F2; --ink-2:#A9BBD0; --muted:#8496AB; --accent:#7FB2E8;
      --good:#5FCB80; --warn:#E0A54A; --ground:#0E141C; --card:#161E29;
      --line:#26313F; --chip:#1D2734;
    }
  }
  :root[data-theme="dark"] {
    --ink:#D8E4F2; --ink-2:#A9BBD0; --muted:#8496AB; --accent:#7FB2E8;
    --good:#5FCB80; --warn:#E0A54A; --ground:#0E141C; --card:#161E29;
    --line:#26313F; --chip:#1D2734;
  }
  :root[data-theme="light"] {
    --ink:#16385E; --ink-2:#42556E; --muted:#6B7A8D; --accent:#2563A8;
    --good:#1A7F37; --warn:#B26A00; --ground:#F2F5F9; --card:#FFFFFF;
    --line:#DFE6EE; --chip:#EEF3F9;
  }

  * { box-sizing:border-box; }
  body {
    margin:0; direction:rtl; background:var(--ground); color:var(--ink-2);
    font-family:-apple-system, BlinkMacSystemFont, "Segoe UI", "Helvetica Neue", Arial, sans-serif;
    font-size:15px; line-height:1.55; -webkit-text-size-adjust:100%;
  }
  .wrap { max-width:760px; margin:0 auto; padding:26px 16px 64px; }

  header { display:flex; flex-direction:column; gap:4px; margin-bottom:18px; }
  h1 { margin:0; font-size:22px; font-weight:700; color:var(--ink); letter-spacing:-.01em; text-wrap:balance; }
  .stamp { font-size:12.5px; color:var(--muted); font-variant-numeric:tabular-nums; }

  .summary { display:flex; gap:8px; flex-wrap:wrap; margin:14px 0 18px; }
  .stat {
    background:var(--card); border:1px solid var(--line); border-radius:10px;
    padding:8px 13px; display:flex; align-items:baseline; gap:7px;
  }
  .stat b { font-size:18px; color:var(--ink); font-variant-numeric:tabular-nums; }
  .stat span { font-size:12.5px; color:var(--muted); }

  .filters { display:flex; gap:7px; flex-wrap:wrap; margin-bottom:16px; }
  .f {
    font:inherit; font-size:13px; cursor:pointer; border:1px solid var(--line);
    background:var(--card); color:var(--ink-2); border-radius:999px; padding:5px 14px;
  }
  .f[aria-pressed="true"] { background:var(--ink); color:var(--ground); border-color:var(--ink); }
  .f:focus-visible { outline:2px solid var(--accent); outline-offset:2px; }

  ol { list-style:none; margin:0; padding:0; display:flex; flex-direction:column; gap:10px; }
  .job {
    background:var(--card); border:1px solid var(--line); border-radius:12px;
    padding:14px 15px; display:grid; grid-template-columns:auto 1fr; gap:0 13px;
  }
  .score {
    grid-row:1 / span 3; width:42px; height:42px; border-radius:11px; display:flex;
    align-items:center; justify-content:center; font-size:18px; font-weight:700;
    color:#fff; background:#7E93AB; font-variant-numeric:tabular-nums;
  }
  .s10, .s9 { background:#1A7F37; } .s8 { background:#2F9E44; }
  .s7 { background:#2563A8; } .s6 { background:#7E93AB; }

  .line1 { display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
  .co { font-weight:700; color:var(--ink); font-size:14.5px; }
  .meta { font-size:12.5px; color:var(--muted); font-variant-numeric:tabular-nums; }
  .pill { font-size:11.5px; font-weight:600; border-radius:999px; padding:2px 9px; background:var(--chip); color:var(--ink-2); }
  .pill.sent { background:rgba(26,127,55,.13); color:var(--good); }
  .title { font-size:15.5px; font-weight:600; color:var(--ink); margin:3px 0 2px; text-wrap:balance; }
  .reason { font-size:13.2px; color:var(--ink-2); margin:0 0 9px; max-width:62ch; }
  a.open {
    font-size:13px; font-weight:600; color:var(--accent); text-decoration:none;
    border-bottom:1px solid color-mix(in srgb, var(--accent) 35%, transparent); padding-bottom:1px;
  }
  a.open:focus-visible { outline:2px solid var(--accent); outline-offset:3px; }

  footer { margin-top:26px; font-size:12.5px; color:var(--muted); border-top:1px solid var(--line); padding-top:14px; }
  @media (max-width:420px) {
    .job { grid-template-columns:auto 1fr; }
    .reason { font-size:13px; }
  }
</style>

<div class="wrap">
  <header>
    <h1>משרות פתוחות – מעקב</h1>
    <div class="stamp">עודכן __STAMP__ · תמונת מצב מתוך jobs.xlsx</div>
  </header>

  <div class="summary">
    <div class="stat"><b>__TOTAL__</b><span>משרות</span></div>
    <div class="stat"><b>__NEW__</b><span>ממתינות</span></div>
    <div class="stat"><b>__SENT__</b><span>הוגשו</span></div>
    <div class="stat"><b>__HIGH__</b><span>ציון 8+</span></div>
  </div>

  <div class="filters">
    <button class="f" data-f="all" aria-pressed="true">הכל</button>
    <button class="f" data-f="new" aria-pressed="false">ממתינות</button>
    <button class="f" data-f="high" aria-pressed="false">ציון 8+</button>
    <button class="f" data-f="north" aria-pressed="false">צפון</button>
  </div>

  <ol id="list"></ol>

  <footer>
    דף קריאה בלבד. יצירת CV, מחיקה ועדכון סטטוס נשארים בדשבורד המקומי במחשב.
    אחרי כל סריקה: <code>python scripts/make_web_report.py</code> ואז פרסום מחדש לאותה כתובת.
  </footer>
</div>

<script>
  const JOBS = __DATA__;
  const NORTH = ["חיפה","יוקנעם","כרמיאל","נצרת","מגדל העמק","מגדל תפן","קיסריה","אור עקיבא","עכו","נהריה","טבריה","עפולה"];
  const isNorth = j => NORTH.some(c => (j.loc || "").includes(c));
  const esc = s => String(s).replace(/[&<>"]/g, c => ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));
  const list = document.getElementById("list");

  function render(filter) {
    const rows = JOBS.filter(j =>
      filter === "new"   ? j.status === "חדש" :
      filter === "high"  ? j.score >= 8 :
      filter === "north" ? isNorth(j) : true);
    list.innerHTML = rows.map(j => `
      <li class="job">
        <div class="score s${j.score}">${j.score}</div>
        <div class="line1">
          <span class="co">${esc(j.co)}</span>
          <span class="meta">${esc(j.loc)} · ${esc(j.date)}</span>
          <span class="pill ${j.status === "הוגש" ? "sent" : ""}">${esc(j.status)}</span>
        </div>
        <div>
          <div class="title">${esc(j.title)}</div>
          <p class="reason">${esc(j.reason)}</p>
          <a class="open" href="${esc(j.link)}" target="_blank" rel="noopener noreferrer">פתח את המשרה ↗</a>
        </div>
      </li>`).join("");
  }

  document.querySelectorAll(".f").forEach(btn => btn.addEventListener("click", () => {
    document.querySelectorAll(".f").forEach(b => b.setAttribute("aria-pressed", String(b === btn)));
    render(btn.dataset.f);
  }));

  render("all");
</script>
"""


def main():
    jobs = read_jobs()
    jobs.sort(key=lambda j: (-(j["score"] or 0), str(j["date"])))
    html = (HTML
            .replace("__DATA__", json.dumps(jobs, ensure_ascii=False))
            .replace("__STAMP__", date.today().isoformat())
            .replace("__TOTAL__", str(len(jobs)))
            .replace("__NEW__", str(sum(1 for j in jobs if j["status"] == u"חדש")))
            .replace("__SENT__", str(sum(1 for j in jobs if j["status"] == u"הוגש")))
            .replace("__HIGH__", str(sum(1 for j in jobs if (j["score"] or 0) >= 8))))
    with io.open(OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"report: {OUT} ({len(jobs)} jobs)")


if __name__ == "__main__":
    main()
