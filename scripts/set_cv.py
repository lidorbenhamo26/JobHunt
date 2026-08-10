# -*- coding: utf-8 -*-
"""Record a generated CV file on a tracker row and rebuild the dashboard.

Usage: python set_cv.py <job_number> <cv_filename_in_output_dir>
"""
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
TRACKER = ROOT / "jobs.xlsx"


def main():
    job_no, cv_name = int(sys.argv[1]), sys.argv[2]
    cv_path = ROOT / "output" / cv_name
    if not cv_path.exists():
        sys.exit(f"CV file not found: {cv_path}")
    wb = load_workbook(TRACKER)
    ws = wb["Jobs"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == job_no:
            cell = row[10]
            # plain filename value only - openpyxl hyperlinks corrupt on later delete_rows
            cell.value = cv_name
            cell.font = Font(name="Arial", size=10, color="2563A8", underline="single")
            wb.save(TRACKER)
            print(f"row {job_no}: CV set -> {cv_name}")
            break
    else:
        sys.exit(f"job #{job_no} not found in tracker")
    subprocess.run([sys.executable, str(ROOT / "scripts" / "make_dashboard.py")], check=True)


if __name__ == "__main__":
    main()
