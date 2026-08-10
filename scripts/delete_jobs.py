# -*- coding: utf-8 -*-
"""Delete jobs from the tracker by their # id and rebuild the dashboard.

Usage: python delete_jobs.py <id> [<id> ...]
IDs of remaining rows are NOT renumbered (stable references).
"""
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TRACKER = ROOT / "jobs.xlsx"


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    ids = {int(a) for a in sys.argv[1:]}
    wb = load_workbook(TRACKER)
    ws = wb["Jobs"]
    to_delete = [row[0].row for row in ws.iter_rows(min_row=2) if row[0].value in ids]
    found = {ws.cell(row=r, column=1).value for r in to_delete}
    if "Deleted" not in wb.sheetnames:
        dl = wb.create_sheet("Deleted")
        dl.sheet_view.rightToLeft = True
        for col, head in enumerate(("תאריך מחיקה", "חברה", "משרה", "קישור", "Job ID"), start=1):
            dl.cell(row=1, column=col, value=head)
    dl = wb["Deleted"]
    from datetime import date
    for r in to_delete:
        row = ws[r]
        link = row[7].hyperlink.target if row[7].hyperlink else row[7].value
        dl.append([date.today().isoformat(), row[2].value, row[3].value, link, row[8].value])
    for r in sorted(to_delete, reverse=True):
        ws.delete_rows(r)
    wb.save(TRACKER)
    missing = ids - found
    print(f"deleted={sorted(found)}" + (f" not_found={sorted(missing)}" if missing else ""))
    subprocess.run([sys.executable, str(ROOT / "scripts" / "make_dashboard.py")], check=True)


if __name__ == "__main__":
    main()
