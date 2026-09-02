# -*- coding: utf-8 -*-
"""Local CV server: dashboard button -> headless Claude Code -> tailored CV PDF.

Runs on http://127.0.0.1:8765 (localhost only).
Endpoints:
  GET /health          -> {"ok": true}
  GET /generate?id=N   -> starts `claude -p "/generate-cv N"` in background
  GET /status?id=N     -> {"status": "running"|"done"|"error", "detail": "..."}
  GET /set-status?id=N&status=S -> updates jobs.xlsx status + rebuilds dashboard
  GET /submit?id=N&mode=review|auto -> starts `codex exec` that fills the
                          application form in the user's Chrome. mode=review
                          (default, "הכנה"): fills the form in its own tab,
                          publishes a summary+screenshot to the dashboard,
                          leaves the tab open and finishes - nothing is sent
                          until the user approves (or submits manually).
                          mode=auto: Codex clicks Submit itself (full-auto).
  GET /submit-status?id=N -> {"status": "running"|"ready"|"login"|"captcha"|
                              "needs_input"|"error", "detail": "..."}
  GET /pending         -> jobs whose Codex run stopped on a missing answer:
                          {"pending": [{"id": N, "q": "<the question>"}]}
  POST /submit-answer  -> body {"id": N, "question": "...", "answer": "..."}:
                          appends the user's answer to application-answers.md
                          and resumes that job's Codex session automatically
  GET /review?id=N     -> field-by-field summary (+screenshot url) Codex
                          published for this job, for remote/phone review
  GET /submit-approve?id=N -> user approved the summary: resume the job's
                          Codex session so it clicks Submit in the open tab
  GET /submissions     -> every submission Codex touched (persisted in
                          output/submissions/<id>.json): status, Codex's last
                          full message, its open QUESTION, parsed review
                          fields, screenshot, chat history - drives the
                          "הגשות לייב" page
  POST /submit-reply   -> body {"id": N, "text": "...", "save_answer": bool}:
                          free-text reply to Codex (answer a QUESTION, ask for
                          a field change, ...) - resumes the job's session
  GET /submit-cancel?id=N  -> drop the submission (nothing is sent)
  GET /submit-dismiss?id=N -> hide a finished record from the live page
  GET /scan?mode=cheap|full -> runs the weekly-job-scan routine NOW in headless
                          Claude (the same SKILL.md the scheduled task uses).
                          cheap = the daily HTTP-only freshness check;
                          full   = every HTTP-reachable source + 7-day window.
                          Headless has no browser/WhatsApp, so a full run here
                          does NOT touch .last-full-scan - the real Sunday scan
                          with browser tools still happens.
  GET /scan-status     -> {"status": "idle"|"running"|"done"|"error", ...}

Start: start-cv-server.bat (or: python scripts\cv_server.py)
Set CV_DRYRUN=1 to test the plumbing without calling Claude/Codex.
"""
import json
import mimetypes
import os
import re
import shutil
import socket
import subprocess
import sys
import threading
import time
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

ROOT = Path(__file__).resolve().parent.parent
PORT = int(os.environ.get("CV_PORT") or 8765)  # CV_PORT: run a test instance beside the real one
TIMEOUT_SEC = 1500  # 25 min per generation


def _contact_email():
    """Applicant email lives in master-cv.json, not in code."""
    try:
        cv = json.loads((ROOT / "master-cv.json").read_text(encoding="utf-8-sig"))
        return cv["contact"]["email"]
    except Exception:  # noqa: BLE001
        return "the account already signed in to this Chrome profile"


ACCOUNT_EMAIL = _contact_email()


def grade_sheet_path():
    pdfs = sorted((ROOT / "transcript").glob("*.pdf"))
    return str(pdfs[0]) if pdfs else "(no grade sheet on file - leave that field empty)"

jobs = {}  # id -> {"status": ..., "detail": ...}
lock = threading.Lock()


def set_status(job_id, status, detail=""):
    with lock:
        jobs[job_id] = {"status": status, "detail": detail}


# ---------------------------------------------------------------- task console
# Live task registry for the dashboard console: every Claude/Codex run is a
# task with a streamed log the UI can tail incrementally.
tasks = {}       # key -> task dict
task_order = []  # keys, oldest first
MAX_TASKS = 20   # history shown in the console
MAX_LOG_LINES = 500


def new_task(kind, job_id, label):
    t = {"key": f"{kind}-{job_id}-{int(time.time())}", "kind": kind,
         "job_id": job_id, "label": label, "status": "running", "detail": "",
         "started": time.time(), "ended": None, "log": [], "base": 0}
    with lock:
        tasks[t["key"]] = t
        task_order.append(t["key"])
        while len(task_order) > MAX_TASKS:
            tasks.pop(task_order.pop(0), None)
    return t


def task_log(t, text):
    with lock:
        for line in text.splitlines():
            line = line.rstrip()
            if line:
                t["log"].append(line[:400])
        drop = len(t["log"]) - MAX_LOG_LINES
        if drop > 0:
            t["log"] = t["log"][drop:]
            t["base"] += drop


def end_task(t, status, detail=""):
    with lock:
        if t.get("stop_requested"):
            status, detail = "stopped", "נעצר ידנית מהקונסול"
        t["status"] = status
        t["detail"] = str(detail)[:500]
        t["ended"] = time.time()
        t.pop("proc", None)


def stop_task(key):
    """Kill a running task's process tree. Returns (ok, detail)."""
    with lock:
        t = tasks.get(key)
        if not t:
            return False, "task not found"
        proc = t.get("proc")
        if t["status"] != "running" or not proc:
            return False, "task not running"
        t["stop_requested"] = True
    try:
        subprocess.run(["taskkill", "/PID", str(proc.pid), "/T", "/F"],
                       capture_output=True, timeout=15,
                       creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
        return True, "stopped"
    except Exception as e:  # noqa: BLE001
        return False, str(e)


def run_streamed(cmd, t, timeout, line_filter=None):
    """Popen with live line capture into the task log.
    Returns (returncode, tail_of_raw_output). Raises TimeoutExpired."""
    proc = subprocess.Popen(
        cmd, cwd=str(ROOT), stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        stdin=subprocess.DEVNULL, text=True, encoding="utf-8", errors="replace",
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )
    with lock:
        t["proc"] = proc
    raw = []

    def reader():
        for line in proc.stdout:
            raw.append(line)
            if len(raw) > 2000:
                del raw[:200]
            shown = line_filter(line) if line_filter else line.rstrip()
            if shown:
                task_log(t, shown)

    th = threading.Thread(target=reader, daemon=True)
    th.start()
    try:
        proc.wait(timeout=timeout)
    except subprocess.TimeoutExpired:
        proc.kill()
        proc.wait(timeout=10)
        raise
    finally:
        th.join(timeout=10)
    return proc.returncode, "".join(raw)


def claude_line_filter(line):
    """Turn `claude -p --output-format stream-json` events into readable log lines."""
    s = line.strip()
    if not s:
        return None
    try:
        ev = json.loads(s)
    except ValueError:
        return s[:300]
    typ = ev.get("type")
    if typ == "system":
        return f"- session {ev.get('subtype', '')} model={ev.get('model', '?')}"
    if typ == "assistant":
        outs = []
        for c in (ev.get("message") or {}).get("content", []):
            if c.get("type") == "text" and (c.get("text") or "").strip():
                outs.append(c["text"].strip()[:200])
            elif c.get("type") == "tool_use":
                inp = c.get("input") or {}
                hint = (inp.get("description") or inp.get("command")
                        or inp.get("file_path") or inp.get("skill") or "")
                outs.append(f"> {c.get('name', 'tool')}: {str(hint)[:150]}")
        return "\n".join(outs) or None
    if typ == "result":
        return f"= finished: {ev.get('subtype', 'done')} ({ev.get('num_turns', '?')} turns)"
    return None


def find_claude():
    cand = shutil.which("claude")
    if cand:
        return cand
    home = Path.home()
    for p in (home / ".local" / "bin" / "claude.exe",
              home / ".local" / "bin" / "claude.cmd",
              home / ".local" / "bin" / "claude",
              home / "AppData" / "Roaming" / "npm" / "claude.cmd"):
        if p.exists():
            return str(p)
    return None


def cv_recorded(job_id):
    """True when the tracker row has a CV filename and that file exists on disk."""
    try:
        from openpyxl import load_workbook
        ws = load_workbook(ROOT / "jobs.xlsx")["Jobs"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == job_id:
                return bool(row[10]) and (ROOT / "output" / str(row[10])).exists()
    except Exception:
        pass
    return False


def run_generate(job_id):
    t = new_task("claude", job_id, f"יצירת CV למשרה #{job_id}")
    line_filter = claude_line_filter
    if os.environ.get("CV_DRYRUN"):
        line_filter = None
        cmd = [sys.executable, "-c", "import time; time.sleep(2); print('dry-run ok')"]
    else:
        claude = find_claude()
        if not claude:
            set_status(job_id, "error", "claude CLI not found (PATH / ~/.local/bin)")
            end_task(t, "error", "claude CLI not found")
            return
        prompt = (f"/generate-cv {job_id} "
                  "(headless run via cv_server: browser tools unavailable, tailor from WebFetch + tracker row; "
                  "no SendUserFile; no questions; run fully autonomously)")
        cmd = [
            claude, "-p", prompt,
            "--permission-mode", "acceptEdits",
            # Agent/Task: the generate-cv skill spawns the CV review subagent
            "--allowedTools",
            "Bash,WebFetch,WebSearch,Read,Write,Edit,Glob,Grep,Skill,TodoWrite,Agent,Task",
            "--output-format", "stream-json", "--verbose",
        ]
    try:
        rc, out = run_streamed(cmd, t, TIMEOUT_SEC, line_filter)
        tail = out.strip()[-2000:]
        ok = cv_recorded(job_id) if not os.environ.get("CV_DRYRUN") else rc == 0
        if ok:
            set_status(job_id, "done", tail)
            end_task(t, "done", "ה-CV נוצר ונרשם בטבלה")
        elif rc == 0:
            msg = f"claude finished but no CV was recorded: {tail[-400:]}"
            set_status(job_id, "error", msg)
            end_task(t, "error", msg)
        elif t.get("stop_requested"):
            set_status(job_id, "error", "נעצר ידנית")
            end_task(t, "error", "")
        else:
            set_status(job_id, "error", f"exit {rc}: {tail}")
            end_task(t, "error", f"claude exited with code {rc}")
    except subprocess.TimeoutExpired:
        set_status(job_id, "error", f"timeout after {TIMEOUT_SEC}s")
        end_task(t, "error", f"timeout after {TIMEOUT_SEC}s")
    except Exception as e:  # noqa: BLE001
        set_status(job_id, "error", str(e))
        end_task(t, "error", str(e))


# ---------------------------------------------------------------- job scan
# "הרץ עכשיו" from the dashboard: the same routine the scheduled task runs,
# started on demand. Headless Claude has no in-app browser and no WhatsApp, so
# both modes are HTTP-only; a full run here therefore does NOT stamp
# .last-full-scan (that stays the scheduled, browser-capable scan's job).
SCAN_TIMEOUT_SEC = 5400  # 90 min ceiling - a full scan touches many sources
SCAN_SKILL = Path.home() / ".claude" / "scheduled-tasks" / "weekly-job-scan" / "SKILL.md"
scan_state = {"status": "idle", "mode": "", "detail": "", "key": "", "started": 0}
scan_lock = threading.Lock()  # one scan at a time


def scan_prompt(mode):
    common = (
        f"Read the scheduled-task file at {SCAN_SKILL} and execute it now, end to end.\n"
        "IMPORTANT CONTEXT FOR THIS RUN:\n"
        "- It was started on demand from the JobHunt dashboard, NOT by the scheduler. "
        "Ignore the cadence check in the file: do not read ROOT\\.last-full-scan to decide the scan type, "
        "it is fixed below.\n"
        "- Headless run: there is no interactive user and NO browser tools and NO WhatsApp. "
        "Use pure HTTP (Bash/python/curl + WebFetch) only. Never ask a question; make reasonable "
        "choices and note them in the output.\n"
        "- Still run step 6 (the end-of-run review subagent: rejection audit + source discovery).\n"
        "- Skip the PushNotification step - the dashboard console shows the result instead.\n"
        "- Finish by printing ONE Hebrew summary line: how many jobs were added and how many scored 8+.\n"
    )
    if mode == "cheap":
        return common + (
            "SCAN TYPE: **CHEAP SCAN**. Follow the cheap-scan rules exactly - NVIDIA/Intel/Amazon APIs, "
            "the LinkedIn guest Haifa-area sweep (max 3 requests) and the JobMaster north query, all with the "
            "72-hour freshness window. Prefilter from list responses; fetch full descriptions only for postings "
            "that pass, cap 6. Do not write .last-full-scan (this is a cheap scan).\n"
        )
    return common + (
        "SCAN TYPE: **FULL SCAN**, with one adjustment for headless mode. Cover every source in SOURCES.md that "
        "is reachable over pure HTTP: NVIDIA, Intel, Amazon, LinkedIn guest (rotate keywords, mind the rate "
        "limits), JobMaster, plus any other HTTP recipe in the file (e.g. the Greenhouse Job Board API under "
        "'Not yet wired'). Use the 7-day freshness window for the full-only sources and 72h for the daily ones. "
        "The browser-only sources (Microsoft, Google, Apple, Check Point, Drushim, AllJobs, WhatsApp) cannot run "
        "headless - list them explicitly at the end as skipped. Update SOURCES.md if an endpoint changed. "
        "DO NOT write ROOT\\.last-full-scan: a headless full scan does not replace the scheduled browser-capable "
        "full scan, and stamping it would suppress the real one.\n"
    )


def run_scan(mode):
    label = "סריקת משרות מהירה" if mode == "cheap" else "סריקת משרות מלאה"
    t = new_task("scan", 0, label)
    with lock:
        scan_state.update({"status": "running", "mode": mode, "detail": "",
                           "key": t["key"], "started": t["started"]})
    line_filter = claude_line_filter
    if os.environ.get("CV_DRYRUN"):
        line_filter = None
        cmd = [sys.executable, "-c", "import time; time.sleep(3); print('dry-run scan ok')"]
    else:
        claude = find_claude()
        if not claude:
            end_task(t, "error", "claude CLI not found")
            with lock:
                scan_state.update({"status": "error", "detail": "claude CLI not found"})
            return
        if not SCAN_SKILL.exists():
            end_task(t, "error", f"scan routine not found at {SCAN_SKILL}")
            with lock:
                scan_state.update({"status": "error", "detail": "SKILL.md not found"})
            return
        cmd = [
            claude, "-p", scan_prompt(mode),
            "--permission-mode", "acceptEdits",
            "--allowedTools",
            "Bash,WebFetch,WebSearch,Read,Write,Edit,Glob,Grep,Agent,Task,Skill,TodoWrite",
            "--output-format", "stream-json", "--verbose",
        ]
    try:
        rc, out = run_streamed(cmd, t, SCAN_TIMEOUT_SEC, line_filter)
        tail = out.strip()[-1500:]
        if rc == 0:
            end_task(t, "done", "הסריקה הסתיימה - רענן את הלוח")
            with lock:
                scan_state.update({"status": "done", "detail": tail[-400:]})
        elif t.get("stop_requested"):
            end_task(t, "error", "")
            with lock:
                scan_state.update({"status": "error", "detail": "נעצר ידנית"})
        else:
            end_task(t, "error", f"claude exited with code {rc}")
            with lock:
                scan_state.update({"status": "error", "detail": f"exit {rc}: {tail[-300:]}"})
    except subprocess.TimeoutExpired:
        end_task(t, "error", f"timeout after {SCAN_TIMEOUT_SEC}s")
        with lock:
            scan_state.update({"status": "error", "detail": f"timeout after {SCAN_TIMEOUT_SEC}s"})
    except Exception as e:  # noqa: BLE001
        end_task(t, "error", str(e))
        with lock:
            scan_state.update({"status": "error", "detail": str(e)})
    finally:
        try:
            scan_lock.release()
        except RuntimeError:
            pass


# ---------------------------------------------------------------- submit (Codex)
SUBMIT_TIMEOUT_SEC = 2400  # 40 min ceiling per Codex form-filling run
subs = {}  # id -> {"status": ..., "detail": ...}
submit_running = threading.Lock()  # one Chrome-driving run at a time


def set_sub(job_id, status, detail=""):
    with lock:
        subs[job_id] = {"status": status, "detail": detail}


def review_file(job_id):
    """Field-by-field summary Codex writes before waiting (review mode)."""
    return ROOT / "output" / f".review_{job_id}.txt"


def shot_file(job_id):
    """Screenshot of the completed form (best effort, review mode)."""
    return ROOT / "output" / f".review_{job_id}.png"


def pending_file(job_id):
    """Question Codex stopped on (NEEDS_INPUT) - drives the answer box on the
    dashboard card. Written when a run ends needs_input, cleared once the form
    is complete (ready/submitted) or a fresh run starts."""
    return ROOT / "output" / f".needs_input_{job_id}.txt"


# ---------------------------------------------------------------- submission records
# One JSON per job under output/submissions/: everything the "הגשות לייב" page
# shows - status, Codex's last full message, its open question, the parsed
# review fields, and the back-and-forth history. Survives server restarts.
SUB_DIR = ROOT / "output" / "submissions"
sub_lock = threading.Lock()


def sub_file(job_id):
    return SUB_DIR / f"{job_id}.json"


def load_sub(job_id):
    try:
        return json.loads(sub_file(job_id).read_text(encoding="utf-8-sig"))
    except Exception:  # noqa: BLE001
        return None


def save_sub(job_id, **fields):
    """Merge fields into the job's submission record (creates it)."""
    with sub_lock:
        rec = load_sub(job_id) or {"id": job_id, "history": []}
        rec.update(fields)
        rec["updated"] = time.time()
        try:
            SUB_DIR.mkdir(parents=True, exist_ok=True)
            sub_file(job_id).write_text(json.dumps(rec, ensure_ascii=False, indent=1),
                                        encoding="utf-8")
        except OSError:
            pass
        return rec


def sub_note(job_id, who, text, kind=""):
    """Append one turn ('codex' or 'me') to the record's history."""
    text = str(text or "").strip()
    if not text:
        return
    with sub_lock:
        rec = load_sub(job_id) or {"id": job_id, "history": []}
        rec.setdefault("history", []).append(
            {"ts": time.time(), "who": who, "kind": kind, "text": text[:6000]})
        rec["history"] = rec["history"][-40:]
        rec["updated"] = time.time()
        try:
            SUB_DIR.mkdir(parents=True, exist_ok=True)
            sub_file(job_id).write_text(json.dumps(rec, ensure_ascii=False, indent=1),
                                        encoding="utf-8")
        except OSError:
            pass


def parse_review_fields(txt):
    """'Field label: value entered' lines -> [{label, value, inferred}]."""
    out = []
    for line in (txt or "").splitlines():
        s = line.strip().lstrip("-*• ").strip()
        if not s or ":" not in s:
            continue
        label, _, value = s.partition(":")
        label, value = label.strip(), value.strip()
        if not label or len(label) > 80:
            continue
        inferred = "[inferred]" in value.lower()
        value = re.sub(r"\s*\[inferred\]\s*", " ", value, flags=re.I).strip()
        out.append({"label": label, "value": value, "inferred": inferred})
    return out


def question_of(status, detail):
    """The text Codex wants answered, without the protocol token."""
    return re.sub(r"^(QUESTION|NEEDS_INPUT|READY_FOR_REVIEW|LOGIN_REQUIRED|FAILED):?\s*",
                  "", str(detail or "")).strip()


ANSWERS_FILE = ROOT / "application-answers.md"
DASH_ANS_HEADER = "## Answers added from the dashboard"
answers_lock = threading.Lock()


def record_answer(job_id, company, question, answer):
    """Append a Q/A the user typed on the dashboard to application-answers.md,
    so this question is answered automatically from now on. Returns
    (ok, detail, raw_line) - raw_line feeds the Claude polish pass."""
    import datetime
    tag = f"[{datetime.date.today().isoformat()} | job #{job_id}" + (
        f" {company}]" if company else "]")
    qa = f"Q: {question} -> A: {answer}" if question else f"A: {answer}"
    line = f"- {tag} {qa}"
    with answers_lock:
        try:
            text = (ANSWERS_FILE.read_text(encoding="utf-8-sig")
                    if ANSWERS_FILE.exists()
                    else "# Application answers - standard form answers (not in the CV)\n")
            if DASH_ANS_HEADER not in text:
                text = text.rstrip("\n") + (
                    f"\n\n{DASH_ANS_HEADER}\n"
                    "Recorded automatically when Lidor answers a NEEDS_INPUT "
                    "question from the dashboard. Answers may be in Hebrew - "
                    "translate to the form's language when filling. A Claude "
                    "pass reviews each entry and promotes it into the proper "
                    "sections above.\n")
            text = text.rstrip("\n") + f"\n{line}\n"
            ANSWERS_FILE.write_text(text, encoding="utf-8")
        except OSError as e:
            return False, str(e), line
    return True, "ok", line


# One polish run at a time - two quick answers must not edit the file at once.
POLISH_TIMEOUT_SEC = 600
polish_lock = threading.Lock()


def run_polish_answer(job_id, company, title, raw_line, question, answer):
    """Headless Claude pass: review the user's raw answer against the question
    and turn it into a proper standing entry in application-answers.md
    (English phrasing, generalized, filed under the right section). Runs in
    the background - Codex already resumed with the raw answer."""
    if os.environ.get("CV_DRYRUN"):
        return
    claude = find_claude()
    if not claude:
        return
    import datetime
    today = datetime.date.today().isoformat()
    prompt = f"""Curate a new standing answer in {ANSWERS_FILE} (edit ONLY that file).

The user just answered a job-application form question that Codex could not answer. The raw entry was appended to the end of the file, exactly:
{raw_line}

Context: job #{job_id}, company: {company or '(unknown)'}, role: {title or '(unknown)'}.
Form question: {question or '(not recorded)'}
User's raw answer (may be Hebrew): {answer}

Do:
1. Review the answer against the question - does it truthfully and fully answer it? Phrase the standing version in clear English (translate faithfully; keep numbers, dates and ranges exactly as given).
2. Generalize it so it is reusable for future applications when the question is generic (drop this job's specifics unless they ARE the answer). If the answer only holds for this specific company/role, keep it scoped and say so in the entry.
3. File it in the most fitting existing section (Identity & contact / Work authorization / Employment status / Standard answers / Grade sheet). Only if none fits, leave it under "{DASH_ANS_HEADER}".
4. Write it as one bullet matching the file's style: "- <Question or label>: <answer>. (added via dashboard {today}, job #{job_id})" - then DELETE the raw line quoted above.
5. If the new answer contradicts an existing entry, keep both and append " | CONFLICT - verify with Lidor" to the new one. Never silently change existing entries.

Rules: never invent facts beyond the user's answer; never use the em-dash character, plain "-" only; touch no other file; no questions - run fully autonomously."""
    t = new_task("claude", job_id, f"ליטוש התשובה החדשה בתשובות הקבועות (#{job_id})")
    cmd = [claude, "-p", prompt,
           "--permission-mode", "acceptEdits",
           "--allowedTools", "Read,Edit,Write",
           "--output-format", "stream-json", "--verbose"]
    with polish_lock:
        try:
            rc, _ = run_streamed(cmd, t, POLISH_TIMEOUT_SEC, claude_line_filter)
            if rc == 0:
                end_task(t, "done", "התשובה נוסחה ושובצה בקובץ התשובות")
            else:
                end_task(t, "error",
                         f"claude exit {rc} - התשובה הגולמית נשארה בסוף הקובץ")
        except subprocess.TimeoutExpired:
            end_task(t, "error", "timeout - התשובה הגולמית נשארה בסוף הקובץ")
        except Exception as e:  # noqa: BLE001
            end_task(t, "error", str(e))


def find_codex():
    base = Path(os.environ.get("LOCALAPPDATA", "")) / "OpenAI" / "Codex" / "bin"
    cands = []
    if base.exists():
        cands += sorted(base.glob("*/codex.exe"),
                        key=lambda p: p.stat().st_mtime, reverse=True)
        if (base / "codex.exe").exists():
            cands.append(base / "codex.exe")
    hit = shutil.which("codex")
    if hit:
        cands.append(Path(hit))
    return str(cands[0]) if cands else None


def get_job_row(job_id):
    from openpyxl import load_workbook
    ws = load_workbook(ROOT / "jobs.xlsx")["Jobs"]
    keys = ["id", "date", "company", "title", "location", "score",
            "reason", "link", "job_id", "status", "cv", "notes"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == job_id:
            vals = [c.hyperlink.target if i == 7 and c.hyperlink else c.value
                    for i, c in enumerate(row)]
            return dict(zip(keys, vals))
    return None


def build_submit_prompt(job, mode="review"):
    cands = list((ROOT / "output" / "cv-data").glob(f"{job['id']}_*.json"))
    comp = re.sub(r"\W", "", str(job["company"] or "")).lower()[:8]
    by_comp = [p for p in cands
               if comp and comp in re.sub(r"\W", "", p.stem).lower()]
    pick = (max(by_comp, key=lambda p: p.stat().st_mtime) if by_comp
            else max(cands, key=lambda p: p.stat().st_mtime) if cands else None)
    tailored_line = (str(pick) if pick
                     else "(no tailored JSON for this job - use master-cv.json only)")
    answers = ROOT / "application-answers.md"
    if mode == "auto":
        intro = ("Fill the application form completely and, once everything is "
                 "verified correct against the data sources, submit it YOURSELF - "
                 "the user has pre-approved autonomous submission for THIS run "
                 "(full-auto mode; they are often away from the computer). "
                 "THIS RUN IS THE ACTION-TIME CONFIRMATION: the user clicked "
                 "'submit' for this specific job on their dashboard after "
                 "reviewing it, so selecting their Google account, transmitting "
                 "the applicant data below and pressing Submit are all already "
                 "approved - do not stop to ask for that confirmation again.")
        step1 = "Open the posting URL in Chrome."
        step7 = ("Everything filled -> re-read the whole form once, verify every "
                 "field against the data sources, then click the final "
                 "Submit/Send/Apply button YOURSELF. Verify the submission actually "
                 "went through (thank-you page, \"application submitted/received\", "
                 "confirmation number) and finish with: SUBMITTED. If you "
                 "deliberately chose not to submit because something looks wrong or "
                 "uncertain, leave the page open on the review step and finish "
                 "with: READY_FOR_REVIEW: <what stopped you>.")
        submit_rule = ("Submitting the application yourself is allowed and "
                       "expected (step 7).")
    else:
        # review mode = הכנה: fill + publish the package, don't wait. The form
        # stays open in its own tab; the user approves from the dashboard
        # (אשר ושלח resumes the session) or submits manually in the tab.
        intro = ("Fill the application form completely, verify it against the "
                 "data sources, then STOP at the final review step and finish "
                 "immediately - the USER approves later from their dashboard "
                 "(review mode). You NEVER click the final Submit button. "
                 "Selecting the user's Google account and entering the applicant "
                 "data below into the form are already approved (the user "
                 "started this run for this job) - no confirmation needed for that.")
        step1 = ("Open the posting URL in a NEW Chrome tab and do all the work "
                 "in that tab. Leave the tab OPEN when you finish - the user "
                 "will return to it.")
        step7 = (
            "Everything filled -> re-read the whole form once, verify every "
            "field against the data sources. Then publish a review package: "
            f"(1) write {review_file(job['id'])} (UTF-8 plain text, overwrite) "
            "with one line per filled field - 'Field label: value entered' - "
            "plus 'Resume uploaded: <filename>', 'Cover letter: <its full "
            "text, or none>'; mark every answer you inferred rather than "
            "copied verbatim with [inferred]. "
            f"(2) best effort: save a screenshot of the completed form page to "
            f"{shot_file(job['id'])} (PNG; full page if you can, otherwise the "
            "visible part; use your browser tooling or a shell utility - if "
            "you can't produce one, skip it and continue). "
            "Then leave this tab OPEN on the final review step WITHOUT "
            "clicking Submit/Send/Apply, do NOT wait for the user, and finish "
            "immediately with: READY_FOR_REVIEW: form ready in its tab, "
            "awaiting dashboard approval.")
        submit_rule = ("NEVER click the final Submit/Send/Apply button - the "
                       "user approves from the dashboard later (review mode). "
                       "Never close the tab with the filled form.")
    return f"""Job application task. Use Chrome browser control (the user's real Chrome). {intro}

JOB
- Company: {job['company']}
- Role: {job['title']}
- Posting URL: {job['link']}
- Requisition/Job ID: {job['job_id'] or '(none)'}

DATA SOURCES (read these files before filling anything):
- {ROOT / 'master-cv.json'} - contact, education, experience. Facts about the applicant come ONLY from these data sources - never invent or embellish.
- {tailored_line} - role-tailored summary + ordered skills (prefer its summary/skills wording).
- {answers} - standard form answers (city, notice period, etc.), if the file exists.
- CV PDF to upload: {ROOT / 'output' / (job['cv'] or '')}

STEPS
1. {step1} Find the Apply flow for this exact role (use the Job ID if the portal needs a search).
2. Login wall - use your judgment, preferring Google SSO:
   - If the portal offers "Sign in with Google" (or the sign-up flow does), use it with the Google account already signed in to this Chrome profile ({ACCOUNT_EMAIL}). This is allowed and pre-approved by the user, including first-time portal account creation via Google SSO and clicking through Google's account chooser / consent screen for basic identity scopes (name, email, profile picture, openid).
   - If Google itself asks for a password or a 2FA code, if the consent screen requests broad scopes (contacts, Drive, Gmail, calendar), or if the portal only offers email+password login with no active session -> leave the page open and finish with: LOGIN_REQUIRED: <portal name>
3. Fill every form field from the data sources. Upload the CV PDF where a resume is requested. If the portal auto-parses the resume, verify every parsed field and fix mistakes. If a previous draft already has an older resume attached, REPLACE it with the CV PDF listed above. If the form asks for a grade sheet / transcript / academic records (גיליון ציונים), upload {grade_sheet_path()} - only into that dedicated field, never as the resume.
4. CAPTCHA -> leave the page open and finish with: CAPTCHA_REQUIRED
5. A REQUIRED question not answered verbatim in the data sources: answer it YOURSELF when the truthful answer is clearly inferable from them (standard eligibility like "are you 18+", years of experience with a technology, willingness/availability questions covered by application-answers.md). Stop with NEEDS_INPUT: <the question> ONLY when you cannot answer truthfully from the data sources - especially salary expectations, visa/work-authorization status, security clearance, or legal declarations that application-answers.md doesn't cover. Optional unanswered fields: leave blank.
6. Cover letter: skip if optional. If required: 3-5 plain professional sentences strictly from the data sources, tailored to the role.
7. {step7}
8. Anything else you need the human for (a confirmation, an ambiguity, a choice between options, a page that looks wrong) -> write your question in plain words and finish with: QUESTION: <the exact question, with the options if any>. The user reads it on their dashboard and replies there; you will be resumed with their answer. Never ask questions any other way - a question that is not on a QUESTION: line is never seen.
9. Anything else fatal -> FAILED: <short reason>

RULES
- The FINAL line of your answer must be exactly one status line: SUBMITTED or READY_FOR_REVIEW or QUESTION: ... or LOGIN_REQUIRED: ... or CAPTCHA_REQUIRED or NEEDS_INPUT: ... or FAILED: ...
- READY_FOR_REVIEW is ONLY for a form that is completely filled AND whose review package (the summary file from step 7) is written. A form that is not filled yet is never READY_FOR_REVIEW - use QUESTION: / LOGIN_REQUIRED: / NEEDS_INPUT: instead.
- Above the final status line, write a short plain-language report of what you did and what state the page is in (2-6 lines) - the user reads it on the dashboard.
- {submit_rule} Never type passwords, 2FA codes, or payment details - not even if a page asks. Google SSO with the already-signed-in Chrome account is allowed (step 2); creating email+password accounts is not.
- Every fact you enter comes ONLY from the data sources. Never invent, guess, or embellish - a submitted application with a wrong fact is worse than one that stops. When unsure about a material answer, prefer READY_FOR_REVIEW over submitting.
- Stick to this one application - no unrelated browsing, no other tabs' data.
"""


CODEX_NOISE = ("ERROR rmcp::transport::worker", "Reading additional input from stdin")


def codex_line_filter(line):
    """Drop unrelated MCP connection spam (ClickUp/Unity from the global codex config)."""
    s = line.rstrip()
    if not s or any(n in s for n in CODEX_NOISE):
        return None
    return s


# Codex session ids per job, so a stopped run (login/CAPTCHA/missing answer)
# can be resumed with full context. Persisted across server restarts.
# Entry: {"sid": <codex session id>, "mode": "review"|"auto"} - the mode sticks
# for resumes. Old files stored a bare sid string (upgraded to review on load).
SESSION_RE = re.compile(r"session id: ([0-9a-f][0-9a-f-]{15,})")
SESSIONS_FILE = ROOT / "output" / ".submit_sessions.json"


def _load_sessions():
    try:
        raw = json.loads(SESSIONS_FILE.read_text(encoding="utf-8-sig"))
        return {int(k): (v if isinstance(v, dict) else {"sid": v, "mode": "review"})
                for k, v in raw.items()}
    except Exception:  # noqa: BLE001
        return {}


sub_sessions = _load_sessions()


def save_session(job_id, sid, mode):
    with lock:
        sub_sessions[job_id] = {"sid": sid, "mode": mode}
        try:
            SESSIONS_FILE.write_text(
                json.dumps({str(k): v for k, v in sub_sessions.items()}),
                encoding="utf-8")
        except OSError:
            pass


def build_continue_prompt(mode="review", job_id=0, answer_note=None):
    if mode == "auto":
        finish = ("Once the form is complete and every field verified against "
                  "the data sources, click the final Submit/Send/Apply button "
                  "YOURSELF (full-auto mode, user pre-approved for this run), "
                  "confirm the submission went through, and report SUBMITTED.")
    else:
        finish = ("Once the form is complete and every field verified against "
                  "the data sources, publish the review package - write "
                  f"{review_file(job_id)} (one line per field: 'Field label: "
                  "value entered', plus resume/cover-letter lines, [inferred] "
                  "marks) and best-effort save a screenshot of the form to "
                  f"{shot_file(job_id)} - then leave the tab OPEN on the final "
                  "review step WITHOUT clicking Submit/Send/Apply and finish "
                  "immediately (do NOT wait) with: READY_FOR_REVIEW: form "
                  "ready in its tab, awaiting dashboard approval.")
    extra = ""
    if answer_note:
        extra = ("\n\nThe answer you were missing was just added to "
                 "application-answers.md (it may be in Hebrew - translate to "
                 "the form's language when filling):\n" + answer_note)
    return f"""The human says they completed the required action in Chrome (login / CAPTCHA solved / a missing answer was added to application-answers.md - re-read that file if your blocker was a missing answer). Re-check the open application page in Chrome and continue exactly where you stopped.{extra}

Current permissions:
- Google SSO with the Chrome profile's signed-in account ({ACCOUNT_EMAIL}) is allowed, including first-time portal account creation and basic identity consent (name, email, profile picture, openid).
- Answer required questions yourself when the truthful answer is clearly inferable from the data sources; stop with NEEDS_INPUT only for salary / visa / security-clearance / legal questions not covered there.
- {finish}
- Still never type passwords, 2FA codes, or payment details; stop for CAPTCHA; never approve broad OAuth scopes (contacts, Drive, Gmail, calendar); never invent facts - when unsure, prefer READY_FOR_REVIEW.

Same finish protocol: the FINAL line of your answer must be exactly one of: SUBMITTED or READY_FOR_REVIEW or QUESTION: ... or LOGIN_REQUIRED: ... or CAPTCHA_REQUIRED or NEEDS_INPUT: ... or FAILED: ... (READY_FOR_REVIEW only once the form is fully filled and the review file is written; anything you need the human for goes on a QUESTION: line, and a short report of what you did goes above the status line.)"""


def build_reply_prompt(mode, job_id, text):
    """The user typed a free-text message on the live-submissions page."""
    if mode == "auto":
        finish = ("This run is in full-auto mode: once the form is complete and "
                  "verified, click the final Submit/Send/Apply button yourself, "
                  "confirm the submission went through, and report SUBMITTED.")
    else:
        finish = ("This run is in review mode: NEVER click the final "
                  "Submit/Send/Apply button, even if the message sounds like an "
                  "approval - the user has a separate Approve button for that. "
                  "When the form is complete, (re)write the review package - "
                  f"{review_file(job_id)} (one line per field: 'Field label: "
                  "value entered', plus resume/cover-letter lines, [inferred] "
                  f"marks) and a best-effort screenshot to {shot_file(job_id)} - "
                  "leave the tab open on the final review step and finish with: "
                  "READY_FOR_REVIEW: form ready in its tab, awaiting dashboard approval.")
    return f"""The user replied to you from their dashboard (may be Hebrew - translate to the form's language when you need to type it):

\"\"\"{text}\"\"\"

Treat it as their answer / instruction: if it answers your question, act on it; if it asks to change or fix something in the form, re-check the open application page in Chrome, make exactly that change, and re-verify the page; if it is a new fact about the applicant, use it only for this form (it is NOT in the data sources). Then continue exactly where you stopped.

Current permissions:
- Google SSO with the Chrome profile's signed-in account ({ACCOUNT_EMAIL}) is allowed, including first-time portal account creation and basic identity consent (name, email, profile picture, openid). Entering the applicant's data into this form is already approved.
- Answer required questions yourself when the truthful answer is clearly inferable from the data sources or from the user's message; stop with NEEDS_INPUT only for salary / visa / security-clearance / legal questions not covered anywhere.
- {finish}
- Still never type passwords, 2FA codes, or payment details; stop for CAPTCHA; never approve broad OAuth scopes (contacts, Drive, Gmail, calendar); never invent facts.

Finish protocol: write 2-6 plain lines on what you did and the page's state, then the FINAL line must be exactly one of: SUBMITTED or READY_FOR_REVIEW or QUESTION: ... or LOGIN_REQUIRED: ... or CAPTCHA_REQUIRED or NEEDS_INPUT: ... or FAILED: ... (anything you need the human for goes on a QUESTION: line.)"""


def build_approve_prompt(job_id):
    return f"""The user reviewed your field-by-field summary on the dashboard and APPROVED submission. In Chrome, verify the application form you filled is still open and intact (re-open the tab if needed; if the form lost its data, refill it from the same data sources as before). Then click the final Submit/Send/Apply button, verify the submission actually went through (thank-you page, "application submitted/received", confirmation number), and finish with: SUBMITTED.
Still never type passwords, 2FA codes, or payment details; stop for CAPTCHA (CAPTCHA_REQUIRED) or a login wall (LOGIN_REQUIRED: <portal>); if something material looks wrong, do NOT submit - describe it and finish with: QUESTION: <what looks wrong and what you need decided>.
Write 2-6 plain lines on what happened (confirmation text / number if any), then the FINAL line must be exactly one of: SUBMITTED or READY_FOR_REVIEW or QUESTION: ... or LOGIN_REQUIRED: ... or CAPTCHA_REQUIRED or NEEDS_INPUT: ... or FAILED: ..."""


STATUS_TOKENS = [
    ("SUBMITTED", "submitted"),
    ("READY_FOR_REVIEW", "ready"),
    ("QUESTION", "question"),
    ("LOGIN_REQUIRED", "login"),
    ("CAPTCHA_REQUIRED", "captcha"),
    ("NEEDS_INPUT", "needs_input"),
    ("FAILED", "error"),
]


def run_submit(job_id, resume=False, mode="review", approve=False, answer_note=None,
               reply_text=None):
    job = get_job_row(job_id)
    if not job:
        return set_sub(job_id, "error", f"job {job_id} not found")
    if not job["cv"] or not (ROOT / "output" / str(job["cv"])).exists():
        return set_sub(job_id, "error", "אין עדיין CV מותאם - קודם ⚡ Generate CV")

    lastmsg = ROOT / "output" / f".submit_last_{job_id}.txt"
    if os.environ.get("CV_DRYRUN"):
        cmd = [sys.executable, "-c",
               "import time; time.sleep(2); print('dry run - nothing filled\\nQUESTION: dry-run question?')"]
        if resume or approve:
            mode = (sub_sessions.get(job_id) or {}).get("mode", mode)
    else:
        codex = find_codex()
        if not codex:
            return set_sub(job_id, "error", "codex.exe not found (OpenAI Codex app installed?)")
        if resume or approve:
            entry = sub_sessions.get(job_id)
            if not entry:
                return set_sub(job_id, "error",
                               "אין ריצת Codex קודמת למשרה הזאת - הרץ 🚀 הגש רגיל")
            mode = entry.get("mode", "review")  # a resume keeps its run's mode
            if approve:
                prompt = build_approve_prompt(job_id)
            elif reply_text:
                prompt = build_reply_prompt(mode, job_id, reply_text)
            else:
                prompt = build_continue_prompt(mode, job_id, answer_note)
            cmd = [codex, "exec", "resume", "--skip-git-repo-check",
                   "-o", str(lastmsg), entry["sid"], prompt]
        else:
            # fresh run: clear review/question leftovers from a previous attempt
            for p in (review_file(job_id), shot_file(job_id), pending_file(job_id)):
                p.unlink(missing_ok=True)
            cmd = [codex, "exec", "--skip-git-repo-check", "-C", str(ROOT),
                   "-o", str(lastmsg), build_submit_prompt(job, mode)]

    if not submit_running.acquire(blocking=False):
        return set_sub(job_id, "busy", "הגשה אחרת רצה כרגע - חכה שתסתיים")
    verb = ("שליחה מאושרת" if approve
            else "תשובה ל-Codex" if reply_text
            else "המשך הגשה" if resume else "הגשת מועמדות")
    if not approve:
        verb += {"auto": " (אוטומטי)", "review": " (הכנה)"}.get(mode, "")
    t = new_task("codex", job_id, f"{verb} למשרה #{job_id} ({job['company']})")
    if not (resume or approve):
        # fresh run -> fresh record (the old conversation is gone with the session)
        with sub_lock:
            sub_file(job_id).unlink(missing_ok=True)
    save_sub(job_id, company=str(job["company"] or ""), title=str(job["title"] or ""),
             link=str(job["link"] or ""), mode=mode, status="running", detail="",
             question="", task_key=t["key"], started=t["started"], dismissed=False)
    if approve:
        sub_note(job_id, "me", "אישרתי - שלח את הטופס", "approve")

    def filt(line):
        m = SESSION_RE.search(line)
        if m:
            save_session(job_id, m.group(1), mode)
        return codex_line_filter(line)
    try:
        lastmsg.unlink(missing_ok=True)
        rc, raw = run_streamed(cmd, t, SUBMIT_TIMEOUT_SEC, filt)
        # prefer the clean final message (-o); fall back to full stdout/stderr
        out = ""
        if lastmsg.exists():
            out = lastmsg.read_text(encoding="utf-8", errors="replace").strip()
            lastmsg.unlink(missing_ok=True)
        if not out:
            out = raw.strip()
        # last matching status token wins
        best = None
        for token, status in STATUS_TOKENS:
            pos = out.rfind(token)
            if pos > (best[0] if best else -1):
                line = out[pos:].splitlines()[0].strip()
                best = (pos, status, line)
        if best:
            status, line = best[1], best[2]
            # "ready" without a review package = Codex asked for something in
            # prose (typically a confirmation) - surface it as a question so the
            # dashboard never shows "approve & send" for an unfilled form
            if status == "ready" and not review_file(job_id).exists():
                status = "question"
                line = "QUESTION: " + (question_of("ready", line) or
                                       "Codex עצר לפני שמילא את הטופס - קרא את ההודעה שלו וענה")
            set_sub(job_id, status, line)
            end_task(t, status, line)
            # the message above the status line is Codex's report to the user
            report = out[:best[0]].strip()[-4000:]
            sub_note(job_id, "codex", (report + "\n" + line).strip(), status)
            save_sub(job_id, status=status, detail=line, last_message=out[-6000:],
                     question=question_of(status, line) if status in
                     ("question", "needs_input", "login", "captcha", "error") else "")
            if status == "needs_input":
                # persist the question so the dashboard shows the answer box
                # even after a reload / server restart
                try:
                    q = re.sub(r"^NEEDS_INPUT:?\s*", "", line).strip()
                    pending_file(job_id).write_text(q, encoding="utf-8")
                except OSError:
                    pass
            elif status in ("submitted", "ready"):
                pending_file(job_id).unlink(missing_ok=True)
            if status == "submitted":
                for p in (review_file(job_id), shot_file(job_id)):
                    p.unlink(missing_ok=True)
                set_job_status(job_id, "הוגש")
        elif t.get("stop_requested"):
            set_sub(job_id, "error", "נעצר ידנית")
            end_task(t, "error", "")
            save_sub(job_id, status="error", detail="נעצר ידנית", last_message=out[-6000:])
            sub_note(job_id, "codex", "(נעצר ידנית)", "error")
        else:
            msg = f"codex exit {rc}, no status line: {out[-400:]}"
            set_sub(job_id, "error", msg)
            end_task(t, "error", msg)
            save_sub(job_id, status="error", detail=msg, last_message=out[-6000:],
                     question="")
            sub_note(job_id, "codex", out[-1500:] or msg, "error")
    except subprocess.TimeoutExpired:
        set_sub(job_id, "error", f"timeout after {SUBMIT_TIMEOUT_SEC}s")
        end_task(t, "error", f"timeout after {SUBMIT_TIMEOUT_SEC}s")
        save_sub(job_id, status="error", detail=f"timeout after {SUBMIT_TIMEOUT_SEC}s")
    except Exception as e:  # noqa: BLE001
        set_sub(job_id, "error", str(e))
        end_task(t, "error", str(e))
        save_sub(job_id, status="error", detail=str(e))
    finally:
        submit_running.release()


def list_submissions():
    """Every persisted record + derived bits the live page needs."""
    out = []
    if not SUB_DIR.exists():
        return out
    for p in sorted(SUB_DIR.glob("*.json")):
        m = re.fullmatch(r"(\d+)\.json", p.name)
        if not m:
            continue
        job_id = int(m.group(1))
        rec = load_sub(job_id)
        if not rec:
            continue
        with lock:
            live = subs.get(job_id, {}).get("status")
        # the in-memory state wins while a run is alive (the file lags a bit)
        if live == "running":
            rec["status"] = "running"
        rv = review_file(job_id)
        if rv.exists():
            try:
                txt = rv.read_text(encoding="utf-8", errors="replace").strip()
                rec["review"] = txt[-8000:]
                rec["fields"] = parse_review_fields(txt)
            except OSError:
                pass
        shot = shot_file(job_id)
        rec["shot"] = (f"output/{shot.name}?t={int(shot.stat().st_mtime)}"
                       if shot.exists() else None)
        rec["resumable"] = job_id in sub_sessions
        out.append(rec)
    out.sort(key=lambda r: -(r.get("updated") or 0))
    return out


# ---------------------------------------------------------------- batch pipeline
# "הגש נבחרות": phase 1 generates a CV for every checked fresh job that lacks one,
# phase 2 chains Codex form-filling runs one at a time. submit_mode "review"
# (default, הכנה): fill each form, publish its review package, leave it open
# in its own Chrome tab and move on - no waiting; the user approves each tab
# later from the dashboard (אשר ושלח). "auto": Codex submits each application
# itself. Either way it stops for passwords/CAPTCHA/questions it can't answer
# truthfully from the data sources.
batch_state = {"status": "idle"}
batch_stop = threading.Event()


def batch_targets(only_ids=None):
    from openpyxl import load_workbook
    ws = load_workbook(ROOT / "jobs.xlsx")["Jobs"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if only_ids is not None and row[0] not in only_ids:
            continue
        if (row[9] or "חדש") != "חדש":
            continue
        cv_ok = bool(row[10]) and (ROOT / "output" / str(row[10])).exists()
        out.append({"id": row[0], "company": row[2] or "", "score": row[5] or 0,
                    "cv": cv_ok})
    out.sort(key=lambda j: -(j["score"] or 0))
    return out


def set_batch(**kw):
    with lock:
        batch_state.update(kw)


def batch_add(name, val):
    with lock:
        batch_state.setdefault(name, []).append(val)


def run_batch(mode, only_ids=None, submit_mode="review"):
    """mode 'cv' = CVs only; 'full' = CVs then chained Codex submissions.
    only_ids: restrict the run to these job ids (checkbox selection).
    submit_mode: 'review' = user clicks each Submit; 'auto' = Codex submits."""
    targets = batch_targets(only_ids)
    set_batch(status="running", mode=mode, phase="cv", total=len(targets),
              cv_done=[], cv_failed=[], submitted=[], ready=[], blocked=[],
              failed=[], current=None, detail="")
    if not targets:
        set_batch(status="done", phase="done",
                  detail="אין משרות מתאימות בסטטוס חדש - אין מה להגיש")
        return
    sel = f" ({len(targets)} נבחרות)" if only_ids is not None else ""
    t = new_task("batch", 0,
                 ("הגש הכל" if mode == "full" else "CV לכולם") + sel)
    need = [j for j in targets if not j["cv"]]
    task_log(t, f"שלב CV: {len(need)} מתוך {len(targets)} צריכות CV")
    for i, j in enumerate(need, 1):
        if batch_stop.is_set():
            break
        set_batch(current=j["id"], phase="cv")
        task_log(t, f"CV {i}/{len(need)}: #{j['id']} {j['company']}")
        run_generate(j["id"])
        with lock:
            ok = jobs.get(j["id"], {}).get("status") == "done"
        batch_add("cv_done" if ok else "cv_failed", j["id"])

    if mode == "full" and not batch_stop.is_set():
        subq = [j for j in batch_targets(only_ids) if j["cv"]]  # re-read: CVs just landed
        task_log(t, f"שלב הגשות: {len(subq)} משרות עם CV מוכן")
        for i, j in enumerate(subq, 1):
            if batch_stop.is_set():
                break
            set_batch(current=j["id"], phase="submit")
            task_log(t, f"הגשה {i}/{len(subq)}: #{j['id']} {j['company']}")
            run_submit(j["id"], mode=submit_mode)
            with lock:
                st = subs.get(j["id"], {}).get("status")
            if st == "submitted":
                batch_add("submitted", j["id"])
            elif st == "ready":
                batch_add("ready", j["id"])
            elif st in ("login", "captcha", "needs_input"):
                batch_add("blocked", {"id": j["id"], "why": st})
            else:
                batch_add("failed", j["id"])

    with lock:
        s = dict(batch_state)
    summary = (f"CV חדשים: {len(s.get('cv_done', []))} | הוגשו: {len(s.get('submitted', []))}"
               f" | ממתינות לקליק שלך: {len(s.get('ready', []))}"
               f" | תקועות (התחברות/שאלה): {len(s.get('blocked', []))}"
               f" | נכשלו: {len(s.get('cv_failed', [])) + len(s.get('failed', []))}")
    stopped = batch_stop.is_set()
    set_batch(status="done", phase="done", current=None, detail=summary)
    end_task(t, "stopped" if stopped else "done", summary)


ALLOWED_STATUSES = {"חדש", "הוגש", "ראיון", "נדחה", "דילגתי"}
xlsx_lock = threading.Lock()


def delete_job(job_id):
    """Delete a tracker row + rebuild dashboard. Returns (ok, detail)."""
    from openpyxl import load_workbook
    with xlsx_lock:
        try:
            wb = load_workbook(ROOT / "jobs.xlsx")
            ws = wb["Jobs"]
            target = None
            for row in ws.iter_rows(min_row=2):
                if row[0].value == job_id:
                    target = row[0].row
                    break
            if target is None:
                return False, f"job {job_id} not found"
            # tombstone: remember link/job_id so the weekly scan never re-adds a deleted job
            trow = ws[target]
            if "Deleted" not in wb.sheetnames:
                dl = wb.create_sheet("Deleted")
                dl.sheet_view.rightToLeft = True
                for col, head in enumerate(("תאריך מחיקה", "חברה", "משרה", "קישור", "Job ID"), start=1):
                    dl.cell(row=1, column=col, value=head)
            dl = wb["Deleted"]
            import datetime
            dl.append([datetime.date.today().isoformat(), trow[2].value, trow[3].value,
                       trow[7].hyperlink.target if trow[7].hyperlink else trow[7].value,
                       trow[8].value])
            ws.delete_rows(target)
            wb.save(ROOT / "jobs.xlsx")
        except PermissionError:
            return False, "jobs.xlsx פתוח באקסל - סגור אותו ונסה שוב"
        except Exception as e:  # noqa: BLE001
            return False, str(e)
    try:
        subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "make_dashboard.py")],
            cwd=str(ROOT), capture_output=True, timeout=60,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception as e:  # noqa: BLE001
        return True, f"deleted, dashboard refresh failed: {e}"
    return True, "ok"


def restore_job(job_id):
    """Move a row from the Archive sheet back to Jobs (status חדש) + rebuild.
    The restore marker in the notes column stops archive_stale from instantly
    re-archiving a row whose posting date is already past the threshold."""
    import datetime
    from openpyxl import load_workbook
    with xlsx_lock:
        try:
            wb = load_workbook(ROOT / "jobs.xlsx")
            if "Archive" not in wb.sheetnames:
                return False, "אין גיליון ארכיון"
            ar = wb["Archive"]
            target = None
            for row in ar.iter_rows(min_row=2):
                if row[0].value == job_id:
                    target = row
                    break
            if target is None:
                return False, f"משרה {job_id} לא נמצאה בארכיון"
            vals = [c.value for c in target[:13]]
            vals += [None] * (13 - len(vals))
            vals[9] = "חדש"
            note = f"שוחזר מהארכיון {datetime.date.today().isoformat()}"
            vals[11] = f"{vals[11]} | {note}" if vals[11] else note
            ws = wb["Jobs"]
            r = ws.max_row
            while r > 1 and all(c.value is None for c in ws[r]):
                r -= 1
            for col, v in enumerate(vals, start=1):
                ws.cell(row=r + 1, column=col, value=v)
            ar.delete_rows(target[0].row)
            wb.save(ROOT / "jobs.xlsx")
        except PermissionError:
            return False, "jobs.xlsx פתוח באקסל - סגור אותו ונסה שוב"
        except Exception as e:  # noqa: BLE001
            return False, str(e)
    try:
        subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "make_dashboard.py")],
            cwd=str(ROOT), capture_output=True, timeout=60,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception as e:  # noqa: BLE001
        return True, f"restored, dashboard refresh failed: {e}"
    return True, "ok"


def _ensure_deleted_sheet(wb):
    if "Deleted" not in wb.sheetnames:
        dl = wb.create_sheet("Deleted")
        dl.sheet_view.rightToLeft = True
        for col, head in enumerate(("תאריך מחיקה", "חברה", "משרה", "קישור", "Job ID"), start=1):
            dl.cell(row=1, column=col, value=head)
    return wb["Deleted"]


def delete_archived(job_id=None):
    """Permanently remove rows from the Archive sheet (all rows when job_id is
    None). Each removed row leaves a tombstone in Deleted so scans never re-add
    it. Returns (ok, detail, count)."""
    import datetime
    from openpyxl import load_workbook
    with xlsx_lock:
        try:
            wb = load_workbook(ROOT / "jobs.xlsx")
            if "Archive" not in wb.sheetnames:
                return False, "אין גיליון ארכיון", 0
            ar = wb["Archive"]
            dl = _ensure_deleted_sheet(wb)
            targets = []
            for row in ar.iter_rows(min_row=2):
                if row[0].value is None:
                    continue
                if job_id is None or row[0].value == job_id:
                    targets.append(row)
            if not targets:
                return False, f"משרה {job_id} לא נמצאה בארכיון" if job_id else "הארכיון ריק", 0
            today = datetime.date.today().isoformat()
            for row in targets:
                dl.append([today, row[2].value, row[3].value, row[7].value, row[8].value])
            for row in sorted(targets, key=lambda r: r[0].row, reverse=True):
                ar.delete_rows(row[0].row)
            wb.save(ROOT / "jobs.xlsx")
        except PermissionError:
            return False, "jobs.xlsx פתוח באקסל - סגור אותו ונסה שוב", 0
        except Exception as e:  # noqa: BLE001
            return False, str(e), 0
    try:
        subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "make_dashboard.py")],
            cwd=str(ROOT), capture_output=True, timeout=60,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception as e:  # noqa: BLE001
        return True, f"deleted, dashboard refresh failed: {e}", len(targets)
    return True, "ok", len(targets)


def set_job_status(job_id, status):
    """Update tracker status + rebuild dashboard. Returns (ok, detail)."""
    from openpyxl import load_workbook
    with xlsx_lock:
        try:
            wb = load_workbook(ROOT / "jobs.xlsx")
            ws = wb["Jobs"]
            hit = False
            for row in ws.iter_rows(min_row=2):
                if row[0].value == job_id:
                    row[9].value = status
                    if status == "הוגש":
                        import datetime
                        # col 13 "הוגש בתאריך" - drives the follow-up nudge on the dashboard
                        ws.cell(row=row[0].row, column=13,
                                value=datetime.date.today().isoformat())
                    hit = True
                    break
            if not hit:
                return False, f"job {job_id} not found"
            wb.save(ROOT / "jobs.xlsx")
        except PermissionError:
            return False, "jobs.xlsx פתוח באקסל - סגור אותו ונסה שוב"
        except Exception as e:  # noqa: BLE001
            return False, str(e)
    try:
        subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "make_dashboard.py")],
            cwd=str(ROOT), capture_output=True, timeout=60,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except Exception as e:  # noqa: BLE001
        return True, f"status saved, dashboard refresh failed: {e}"
    return True, "ok"


# Pages served over HTTP so phones/other devices can load the dashboard
# (opening dashboard.html from file:// still works as before).
STATIC_PAGES = {"dashboard.html", "archive.html", "profile.html", "report.html",
                "submissions.html"}


class Handler(BaseHTTPRequestHandler):
    def _serve_file(self, rel):
        path = (ROOT / rel).resolve()
        out_dir = (ROOT / "output").resolve()
        allowed = ((path.parent == ROOT and path.name in STATIC_PAGES)
                   or out_dir in path.parents)
        if not allowed or not path.is_file():
            return self._json(404, {"error": "not found"})
        ctype = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        if ctype.startswith("text/") or ctype == "image/svg+xml":
            ctype += "; charset=utf-8"
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def _json(self, code, obj):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        u = urlparse(self.path)
        rel = unquote(u.path).lstrip("/")
        if u.path == "/":
            return self._serve_file("dashboard.html")
        if rel in STATIC_PAGES or rel.startswith("output/"):
            return self._serve_file(rel)

        if u.path == "/health":
            return self._json(200, {"ok": True})

        if u.path == "/tasks":
            with lock:
                out = []
                for k in reversed(task_order):
                    t = tasks.get(k)
                    if t:
                        out.append({"key": t["key"], "kind": t["kind"],
                                    "job_id": t["job_id"], "label": t["label"],
                                    "status": t["status"], "detail": t["detail"],
                                    "started": t["started"], "ended": t["ended"],
                                    "log_len": t["base"] + len(t["log"])})
            try:
                dash_mtime = os.path.getmtime(ROOT / "dashboard.html")
            except OSError:
                dash_mtime = 0
            return self._json(200, {"tasks": out, "now": time.time(),
                                    "dash_mtime": dash_mtime})

        if u.path == "/tasks-clear":
            with lock:
                keep = [k for k in task_order
                        if tasks.get(k, {}).get("status") == "running"]
                removed = len(task_order) - len(keep)
                for k in list(tasks):
                    if k not in keep:
                        tasks.pop(k, None)
                task_order[:] = keep
            return self._json(200, {"ok": True, "removed": removed})

        if u.path == "/task-stop":
            key = parse_qs(u.query).get("key", [""])[0]
            ok, detail = stop_task(key)
            return self._json(200 if ok else 400, {"ok": ok, "detail": detail})

        if u.path == "/task-log":
            key = parse_qs(u.query).get("key", [""])[0]
            frm_raw = parse_qs(u.query).get("from", ["0"])[0]
            frm = int(frm_raw) if frm_raw.isdigit() else 0
            with lock:
                t = tasks.get(key)
                if not t:
                    return self._json(404, {"error": "no such task"})
                start = max(0, frm - t["base"])
                return self._json(200, {"lines": t["log"][start:],
                                        "next": t["base"] + len(t["log"]),
                                        "status": t["status"],
                                        "detail": t["detail"]})

        if u.path == "/batch":
            q = parse_qs(u.query)
            mode = q.get("mode", ["full"])[0]
            if mode not in ("cv", "full"):
                return self._json(400, {"error": "bad mode"})
            submit_mode = q.get("submit_mode", ["review"])[0]
            if submit_mode not in ("review", "auto"):
                return self._json(400, {"error": "bad submit_mode"})
            ids_raw = parse_qs(u.query).get("ids", [""])[0]
            only_ids = None
            if ids_raw:
                try:
                    only_ids = {int(x) for x in ids_raw.split(",") if x.strip()}
                except ValueError:
                    return self._json(400, {"error": "bad ids"})
            # selection is mandatory (user decision 2026-08-10): no blanket runs over
            # every fresh job - CV and submit both act only on checked cards
            if not only_ids:
                return self._json(400, {"error": "no ids",
                                        "detail": "סמן משרות בלוח קודם - הרצה על הכל בוטלה"})
            with lock:
                if batch_state.get("status") == "running":
                    return self._json(200, {"status": "running"})
            batch_stop.clear()
            threading.Thread(target=run_batch, args=(mode, only_ids, submit_mode),
                             daemon=True).start()
            return self._json(200, {"status": "running"})

        if u.path == "/batch-status":
            with lock:
                return self._json(200, dict(batch_state))

        if u.path == "/batch-stop":
            batch_stop.set()
            return self._json(200, {"ok": True, "detail": "נעצר אחרי המשימה הנוכחית"})

        if u.path == "/archive-clear":
            ok, detail, n = delete_archived(None)
            return self._json(200 if ok else 500,
                              {"ok": ok, "detail": detail, "deleted": n})

        if u.path == "/scan":
            mode = parse_qs(u.query).get("mode", ["cheap"])[0]
            if mode not in ("cheap", "full"):
                return self._json(400, {"error": "bad mode"})
            if not scan_lock.acquire(blocking=False):
                with lock:
                    return self._json(200, {"status": "running",
                                            "mode": scan_state.get("mode", ""),
                                            "detail": "סריקה כבר רצה"})
            threading.Thread(target=run_scan, args=(mode,), daemon=True).start()
            return self._json(200, {"status": "running", "mode": mode})

        if u.path == "/scan-status":
            with lock:
                return self._json(200, dict(scan_state))

        if u.path == "/submissions":
            recs = list_submissions()
            attn = sum(1 for r in recs if not r.get("dismissed") and
                       r.get("status") in ("question", "ready", "login", "captcha",
                                           "needs_input"))
            return self._json(200, {"submissions": recs, "attention": attn,
                                    "busy": submit_running.locked(), "now": time.time()})

        if u.path == "/pending":
            out = []
            for p in sorted((ROOT / "output").glob(".needs_input_*.txt")):
                m = re.fullmatch(r"\.needs_input_(\d+)\.txt", p.name)
                if not m:
                    continue
                try:
                    q = p.read_text(encoding="utf-8", errors="replace").strip()
                except OSError:
                    continue
                out.append({"id": int(m.group(1)), "q": q[:500]})
            return self._json(200, {"pending": out})

        raw = parse_qs(u.query).get("id", [""])[0]
        if not re.fullmatch(r"\d+", raw):
            return self._json(400, {"error": "bad id"})
        job_id = int(raw)

        if u.path == "/generate":
            with lock:
                cur = jobs.get(job_id)
                if cur and cur["status"] == "running":
                    return self._json(200, {"status": "running"})
                jobs[job_id] = {"status": "running", "detail": ""}
            threading.Thread(target=run_generate, args=(job_id,), daemon=True).start()
            return self._json(200, {"status": "running"})

        if u.path == "/status":
            with lock:
                return self._json(200, jobs.get(job_id, {"status": "unknown"}))

        if u.path == "/submit":
            mode = parse_qs(u.query).get("mode", ["review"])[0]
            if mode not in ("review", "auto"):
                return self._json(400, {"error": "bad mode"})
            with lock:
                cur = subs.get(job_id)
                if cur and cur["status"] == "running":
                    return self._json(200, {"status": "running"})
                subs[job_id] = {"status": "running", "detail": ""}
            threading.Thread(target=run_submit, args=(job_id, False, mode),
                             daemon=True).start()
            return self._json(200, {"status": "running"})

        if u.path == "/submit-continue":
            with lock:
                cur = subs.get(job_id)
                if cur and cur["status"] == "running":
                    return self._json(200, {"status": "running"})
                subs[job_id] = {"status": "running", "detail": ""}
            threading.Thread(target=run_submit, args=(job_id, True), daemon=True).start()
            return self._json(200, {"status": "running"})

        if u.path == "/submit-status":
            with lock:
                return self._json(200, subs.get(job_id, {"status": "unknown"}))

        if u.path == "/review":
            p = review_file(job_id)
            if not p.exists():
                return self._json(404, {"error": "no review yet"})
            try:
                txt = p.read_text(encoding="utf-8", errors="replace").strip()
            except OSError as e:
                return self._json(500, {"error": str(e)})
            shot = shot_file(job_id)
            return self._json(200, {
                "summary": txt[-8000:],
                "shot": f"output/{shot.name}?t={int(shot.stat().st_mtime)}"
                        if shot.exists() else None,
            })

        if u.path == "/submit-approve":
            with lock:
                alive = subs.get(job_id, {}).get("status") == "running"
            if alive:
                return self._json(200, {"ok": False,
                                        "detail": "Codex עדיין עובד על המשרה - חכה שיסיים"})
            if not sub_sessions.get(job_id):
                return self._json(400, {"ok": False,
                                        "detail": "אין ריצת Codex לאשר - הרץ הגשה קודם"})
            with lock:
                subs[job_id] = {"status": "running", "detail": ""}
            threading.Thread(target=run_submit, args=(job_id,),
                             kwargs={"approve": True}, daemon=True).start()
            return self._json(200, {"ok": True, "via": "resume"})

        if u.path == "/submit-cancel":
            with lock:
                alive = subs.get(job_id, {}).get("status") == "running"
            if alive:
                return self._json(200, {"ok": False,
                                        "detail": "Codex עדיין רץ - עצור אותו מהקונסול קודם"})
            for p in (review_file(job_id), shot_file(job_id), pending_file(job_id)):
                p.unlink(missing_ok=True)
            with lock:
                subs.pop(job_id, None)
            save_sub(job_id, status="cancelled", detail="בוטל מהדשבורד", question="")
            sub_note(job_id, "me", "ביטלתי את ההגשה (הטופס לא נשלח)", "cancel")
            return self._json(200, {"ok": True})

        if u.path == "/submit-dismiss":
            if not load_sub(job_id):
                return self._json(404, {"ok": False, "detail": "no record"})
            save_sub(job_id, dismissed=True)
            return self._json(200, {"ok": True})

        if u.path == "/set-status":
            status = parse_qs(u.query).get("status", [""])[0]
            if status not in ALLOWED_STATUSES:
                return self._json(400, {"ok": False, "detail": "bad status"})
            ok, detail = set_job_status(job_id, status)
            return self._json(200 if ok else 500, {"ok": ok, "detail": detail})

        if u.path == "/delete":
            ok, detail = delete_job(job_id)
            return self._json(200 if ok else 500, {"ok": ok, "detail": detail})

        if u.path == "/restore":
            ok, detail = restore_job(job_id)
            return self._json(200 if ok else 500, {"ok": ok, "detail": detail})

        if u.path == "/archive-delete":
            ok, detail, n = delete_archived(job_id)
            return self._json(200 if ok else 500,
                              {"ok": ok, "detail": detail, "deleted": n})

        return self._json(404, {"error": "not found"})

    def do_POST(self):
        u = urlparse(self.path)
        if u.path not in ("/submit-answer", "/submit-reply"):
            return self._json(404, {"error": "not found"})
        try:
            n = int(self.headers.get("Content-Length") or 0)
            if not 0 < n <= 20000:
                return self._json(400, {"error": "bad length"})
            data = json.loads(self.rfile.read(n).decode("utf-8", "replace"))
            job_id = int(data["id"])
            question = str(data.get("question") or "").strip()
            answer = str(data.get("answer") or data.get("text") or "").strip()
            save_answer = bool(data.get("save_answer"))
        except (ValueError, KeyError, TypeError):
            return self._json(400, {"error": "bad request"})
        if not answer:
            return self._json(400, {"ok": False, "detail": "תשובה ריקה"})
        with lock:
            if subs.get(job_id, {}).get("status") == "running":
                return self._json(200, {"ok": False,
                                        "detail": "Codex עדיין רץ על המשרה - חכה שיסיים"})
        try:
            job = get_job_row(job_id)
        except Exception:  # noqa: BLE001
            job = None
        company = (job or {}).get("company") or ""

        if u.path == "/submit-reply":
            # free-text message to Codex (answer to its QUESTION, a fix request,
            # ...). Not a standing answer unless the user ticked "save".
            if not sub_sessions.get(job_id):
                return self._json(400, {"ok": False,
                                        "detail": "אין ריצת Codex להמשיך - הרץ הגשה קודם"})
            if submit_running.locked():
                return self._json(200, {"ok": False,
                                        "detail": "הגשה אחרת רצה כרגע - נסה שוב כשתסתיים"})
            sub_note(job_id, "me", answer, "reply")
            rec = load_sub(job_id) or {}
            if save_answer:
                q = rec.get("question") or question
                ok, detail, raw_line = record_answer(job_id, company, q, answer)
                if ok:
                    threading.Thread(target=run_polish_answer,
                                     args=(job_id, company, (job or {}).get("title") or "",
                                           raw_line, q, answer), daemon=True).start()
            pending_file(job_id).unlink(missing_ok=True)
            with lock:
                subs[job_id] = {"status": "running", "detail": ""}
            threading.Thread(target=run_submit, args=(job_id, True),
                             kwargs={"reply_text": answer}, daemon=True).start()
            return self._json(200, {"ok": True, "resumed": True})

        ok, detail, raw_line = record_answer(job_id, company, question, answer)
        if not ok:
            return self._json(500, {"ok": False, "detail": detail})
        sub_note(job_id, "me", (f"{question}\n-> " if question else "") + answer, "answer")
        pending_file(job_id).unlink(missing_ok=True)
        # resume the same Codex session with the fresh answer, unless another
        # Chrome-driving run holds the lock (then the card's המשך button works)
        resumed = False
        if sub_sessions.get(job_id) and not submit_running.locked():
            with lock:
                subs[job_id] = {"status": "running", "detail": ""}
            note = f"Q: {question}\nA: {answer}" if question else f"A: {answer}"
            threading.Thread(target=run_submit, args=(job_id, True),
                             kwargs={"answer_note": note}, daemon=True).start()
            resumed = True
        # background Claude pass: review the raw answer against the question
        # and file a polished standing entry (Codex is not blocked by this)
        threading.Thread(target=run_polish_answer,
                         args=(job_id, company, (job or {}).get("title") or "",
                               raw_line, question, answer),
                         daemon=True).start()
        return self._json(200, {"ok": True, "resumed": resumed})

    def log_message(self, fmt, *args):  # quiet
        pass


def lan_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except OSError:
        return None


def main():
    # A second instance must fail loudly, not silently stack on the same port
    # (Windows SO_REUSEADDR allows multiple binds otherwise).
    ThreadingHTTPServer.allow_reuse_address = False
    # 0.0.0.0 so phones / other devices (LAN or Tailscale) can reach the
    # dashboard. Set CV_BIND=127.0.0.1 to go back to localhost-only.
    host = os.environ.get("CV_BIND", "0.0.0.0")
    try:
        server = ThreadingHTTPServer((host, PORT), Handler)
    except OSError:
        print(f"port {PORT} busy - CV server already running")
        return
    mode = "DRY-RUN" if os.environ.get("CV_DRYRUN") else "claude headless"
    ip = lan_ip()
    remote = f"  |  from phone/LAN: http://{ip}:{PORT}" if ip and host != "127.0.0.1" else ""
    print(f"CV server: http://127.0.0.1:{PORT}{remote}  root={ROOT}  mode={mode}")
    server.serve_forever()


if __name__ == "__main__":
    main()
